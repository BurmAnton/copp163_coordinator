import math
from datetime import date, datetime, tzinfo

import pytz
from django.core.exceptions import ObjectDoesNotExist
from django.utils import timezone
from openpyxl import load_workbook

from citizens.models import Citizen, School
from education_centers.models import (Competence, EducationCenter,
                                      EducationProgram, Workshop)
from users.models import User

from .models import (Application, CitizenCategory, Contract, EducationCenterProjectYear,
                     FlowStatus, Group, ProfField, Profstandart, ProgramModule, ProjectYear, Subject)

CONTR_TYPE_CHOICES = {
     "Трёхсторонний с работодателем": 'THR',
     "Трёхсторонний с ЦЗН": 'CZN',
     "Двусторонний": 'SELF'
}
EDUCATION_CHOICES = {
    "Высшее образование – бакалавриат": 'SPVO',
    "Среднее профессиональное образование - техникум, колледж": 'SPVO',
    "Высшее образование – специалитет, магистратура": 'SPVO',
    "Высшее образование – подготовка кадров высшей квалификации": 'SPVO',
    "Среднее общее образование - 11 классов": 'SCHL',
    "Основное общее образование - 9 классов": 'SCHL',
}

def get_sheet(form):
    workbook = load_workbook(form.cleaned_data['import_file'])
    return workbook.active

def cheak_col_match(sheet, fields_names_set):
    i = 0
    col_count = sheet.max_column
    sheet_fields = []
    sheet_col = {}
    if sheet["A2"].value is None:
        return ['Import', 'EmptySheet']
    try:
        for col_header in range(1, col_count+1):
            if sheet.cell(row=1,column=col_header).value is not None:
                sheet_fields.append(sheet.cell(row=1,column=col_header).value)
                sheet_col[col_header] = sheet.cell(row=1,column=col_header).value
        if missing_fields := [
            field for field in fields_names_set if field not in sheet_fields
        ]:
            return ['Error', 'MissingFieldsError', missing_fields]
    except IndexError:
            return ['Error', 'IndexError']
    return [True, sheet_col]

def load_worksheet_dict(sheet, fields_names_set):
    row_count = sheet.max_row
    sheet_dict = {}
    for col in fields_names_set:
        sheet_dict[fields_names_set[col]] = []
        for row in range(2, row_count+1): 
            snils = sheet[f"A{row}"].value
            if snils != None:
                cell_value = sheet.cell(row=row,column=col).value
                try: cell_value = str(math.floor(cell_value))
                except (ValueError, TypeError): pass
                if cell_value is None:
                    cell_value = ""
                sheet_dict[fields_names_set[col]].append(cell_value)
    return sheet_dict


def competencies(program: ProgramModule, form) -> dict:
    try: sheet = get_sheet(form)
    except IndexError: return ['Error', 'IndexError']

    #Требуемые поля таблицы
    fields_names = {
        "Тема/аттестация",
        "Модуль",
        "Тип поля",
        "Форма аттестации",
        "Лекции (ак. часов)",
        "Лекции (содержание)",
        "Практика (ак. часов)",	
        "Практика (содержание)",
        "Консультации (ак. часов)",
        "Консультации (содержание)",
        "Сам. работа (ак. часов)",
        "Сам. работа (содержание)"
    }

    cheak_col_names = cheak_col_match(sheet, fields_names)
    if cheak_col_names[0] != True:
        return cheak_col_names

    sheet = load_worksheet_dict(sheet, cheak_col_names[1])
    modules_list = create_modules(sheet, program)
    modules_dict = { key: 0 for key in modules_list }
    subjects = []

    for row in range(len(sheet['Тип поля'])):
        field_type = sheet["Тип поля"][row]
        if sheet["Модуль"][row] != "":
            module = program.modules.get(name=sheet["Модуль"][row])
        lections_duration, practice_duration, consultations_duration, independent_duration = get_duration(sheet, row)
        if field_type == "Тема":
            modules_dict[module.name] += 1
            subjects.append(
                Subject(
                    module=module,
                    index=modules_dict[module.name],
                    name=sheet["Тема/аттестация"][row],
                    attest_form=sheet["Форма аттестации"][row],
                    lections_duration=lections_duration,
                    practice_duration=practice_duration,
                    consultations_duration=consultations_duration,
                    independent_duration=independent_duration,
                    lections=sheet["Лекции (содержание)"][row],
                    practice=sheet["Практика (содержание)"][row],
                    consultations=sheet["Консультации (содержание)"][row],
                    independent=sheet["Сам. работа (содержание)"][row],
                )
            )
        elif field_type == "Промежуточная аттестация":
            module.attest_form=sheet["Форма аттестации"][row]
            module.lections_duration=lections_duration
            module.practice_duration=practice_duration
            module.consultations_duration=consultations_duration
            module.independent_duration=independent_duration
            module.save()
        elif field_type == "Итоговая аттестация":
            program.exam_attest_form=sheet["Форма аттестации"][row]
            program.exam_lections_duration=lections_duration
            program.exam_practice_duration=practice_duration
            program.exam_consultations_duration=consultations_duration
            program.exam_independent_duration=independent_duration
            program.save()
    
    subjects = Subject.objects.bulk_create(subjects)
    return ['OK',]


def get_duration(sheet: list, row: int) -> list:
    if sheet["Лекции (ак. часов)"][row].isnumeric():
        lections_duration = int(sheet["Лекции (ак. часов)"][row])
    else: lections_duration = 0
    if sheet["Практика (ак. часов)"][row].isnumeric():
        practice_duration = int(sheet["Практика (ак. часов)"][row])
    else: practice_duration = 0
    if sheet["Консультации (ак. часов)"][row].isnumeric():
        consultations_duration = int(sheet["Консультации (ак. часов)"][row])
    else: consultations_duration = 0
    if sheet["Сам. работа (ак. часов)"][row].isnumeric():
        independent_duration = int(sheet["Сам. работа (ак. часов)"][row])
    else: independent_duration = 0

    return [lections_duration, practice_duration, consultations_duration, independent_duration]


def create_modules(sheet: list, program: ProgramModule) -> list:
    modules = []
    modules_names = []
    index = 1
    for name in sheet['Модуль']:
        if name != "" and name not in modules_names:
            modules_names.append(name)
            modules.append(ProgramModule(program=program, name=name, index=index))
            index += 1
    modules = ProgramModule.objects.bulk_create(modules)
    return list(modules_names)


def profstandarts(form):
    try: sheet = get_sheet(form)
    except IndexError: return ['Error', 'IndexError']

    #Требуемые поля таблицы
    fields_names = {
        "Код профессионального стандарта",
        "Область профессиональной деятельности",
        "Вид профессиональной деятельности",
        "Наименование профессионального стандарта",
        "Приказ Минтруда (номер)",
        "Приказ Минтруда (дата)",
        "Регистрационный номер Минюста",	
        "Дата регистрации в Минюсте"
    }
    cheak_col_names = cheak_col_match(sheet, fields_names)
    if cheak_col_names[0] != True:
        return cheak_col_names

    sheet = load_worksheet_dict(sheet, cheak_col_names[1])
    for row in range(len(sheet['Код профессионального стандарта'])):
        code = sheet["Код профессионального стандарта"][row]
        code_field, code_standart = code.split('.')
        #ProfField
        prof_field, _ = ProfField.objects.get_or_create(code=code_field)
        prof_field.name = sheet["Область профессиональной деятельности"][row]
        prof_field.save()
        #Profstandart
        standart, _ = Profstandart.objects.get_or_create(prof_field=prof_field, code=code_standart)
        standart.name = sheet["Наименование профессионального стандарта"][row]
        standart.prof_activity_type = sheet["Вид профессиональной деятельности"][row]
        standart.mintrud_order_number = sheet["Приказ Минтруда (номер)"][row]
        standart.mintrud_order_date = sheet["Приказ Минтруда (дата)"][row]
        standart.minust_order_date = sheet["Регистрационный номер Минюста"][row]
        standart.minust_order_number = sheet["Дата регистрации в Минюсте"][row]
        standart.save()
    return ['OK',]


def import_applications(form, year):
    try: sheet = get_sheet(form)
    except IndexError: return ['Error', 'IndexError']

    #Требуемые поля таблицы
    fields_names = {
        "Номер заявки", "Фамилия", "Имя", "Отчество", "СНИЛС", "Телефон",
        "Пол гражданина","Уровень образования с портала РР", "Email",
        "Регион", "Статус заявки", "Дата одобрения ЦЗН", "Категория (полное)",
        "Тип договора", "Образовательный партнёр", "Начало обучения",
        "Окончание обучения", "Дата создания заявки", "Идентификатор потока",
        "Программа", "Идентификатор образовательной программы", 
        "Дата рождения гражданина", "Срок истечения заявки",
        "Дата заключения договора со слушателем", 
        "Стоимость договора гражданина", "Документ о занятости",
        "Номер договора на организацию обучения"
    }
    cheak_col_names = cheak_col_match(sheet, fields_names)
    if cheak_col_names[0] != True:
        return cheak_col_names

    sheet = load_worksheet_dict(sheet, cheak_col_names[1])
    missing_fields = []
    citizen_added = 0
    citizen_updated = 0
    application_added = 0
    application_updated = 0
    group_added = 0
    group_updated = 0
    project_year = ProjectYear.objects.get(year=year)
    project_year.appls_last_update = datetime.now()
    project_year.save()

    for row in range(len(sheet['Номер заявки'])):
        citizen_input = create_citizen(sheet, row)
        if citizen_input['status'] == True:
            citizen = citizen_input['value']
            if citizen_input['is_new']: citizen_added += 1
            else: citizen_updated += 1
            application_input = create_application(sheet, row, citizen, project_year)
            if application_input['status'] == True:
                if application_input['is_new']: application_added += 1
                else: application_updated += 1
                if sheet["Идентификатор потока"][row] not in ["", None]:
                    application = application_input['value']
                    group_input = create_group(sheet, row, application)
                    if group_input['is_new']: group_added += 1
                    else: group_updated += 1
            elif application_input not in missing_fields:
                if application_input['status'] != 'EdProgramMissingFed':
                    missing_fields.append(application_input)
                elif citizen_input['is_new']: citizen_added -= 1
                else: citizen_updated -= 1
        elif citizen_input not in missing_fields:
            missing_fields.append(citizen_input)
    missing_fields = sorted(missing_fields, key=lambda d: d['status'])
    return ['OK', missing_fields, citizen_added, citizen_updated,
             application_added, application_updated, group_added, group_updated]

def create_citizen(sheet, row):
    snils_number = sheet["СНИЛС"][row]
    if snils_number == "" or snils_number is None:
        return {"status": "SnilsMissing", "value": row+2}
    citizen, is_new = Citizen.objects.get_or_create(
        snils_number=snils_number
    )
    citizen.last_name = sheet["Фамилия"][row]
    citizen.first_name = sheet["Имя"][row]
    citizen.middle_name = sheet["Отчество"][row]
    citizen.phone_number = sheet["Телефон"][row]
    citizen.email = sheet["Email"][row]
    citizen.res_region = sheet["Регион"][row]
    citizen.birthday = datetime.strptime(sheet["Дата рождения гражданина"][row], "%d.%m.%Y")
    if sheet["Пол гражданина"][row] == "Мужской": citizen.sex = 'M'
    elif sheet["Пол гражданина"][row] == "Женский": citizen.sex = 'F'
    education_type = sheet["Уровень образования с портала РР"][row]
    if education_type in EDUCATION_CHOICES:
        citizen.education_type = EDUCATION_CHOICES[education_type]
    citizen.save()

    return {"status": True, "value": citizen, "is_new": is_new}

def create_application(sheet, row, citizen, project_year):
    flow_id = sheet["Номер заявки"][row].replace('=HYPERLINK("https://flow.firpo.info/Requests/Card/', '')
    flow_id = sheet["Номер заявки"][row].replace(')', '')
    flow_id = int(flow_id.split()[1])
    application, is_new = Application.objects.get_or_create(
        flow_id=flow_id,
        project_year=project_year,
        applicant=citizen
    )
    flow_status = sheet["Статус заявки"][row]
    try:
        application.flow_status = FlowStatus.objects.get(off_name=flow_status)
    except ObjectDoesNotExist:
        return {"status": "FlowStatusMissing", "value": flow_status}
    application.creation_date = datetime.strptime(sheet["Дата создания заявки"][row], "%d.%m.%Y")
    if sheet["Срок истечения заявки"][row] != None:
        application.expiration_date = datetime.strptime(sheet["Срок истечения заявки"][row], "%d.%m.%Y")
    if sheet["Дата одобрения ЦЗН"][row] != None:
        csn_prv_date = datetime.strptime(sheet["Дата одобрения ЦЗН"][row], "%d.%m.%Y")
        if csn_prv_date == "": csn_prv_date = None
        application.csn_prv_date = csn_prv_date
    citizen_category = sheet["Категория (полное)"][row]
    citizen_category = CitizenCategory.objects.filter(
            official_name=citizen_category, project_year=project_year)
    if len(citizen_category) == 0:
        application.delete()
        return {"status": "CategoryMissing", "value": sheet["Категория (полное)"][row]}
    application.citizen_category = citizen_category[0]
    contract_type = sheet["Тип договора"][row]
    if contract_type in CONTR_TYPE_CHOICES:
        application.contract_type = CONTR_TYPE_CHOICES[contract_type]
        if sheet["Дата заключения договора со слушателем"][row] != None:
            application.contract_date = datetime.strptime(sheet[
                "Дата заключения договора со слушателем"][row], "%d.%m.%Y")
    else: application.contract_type = "–"

    if sheet["Документ о занятости"][row] in ["Не проверен", "Подтвержден"]:
        application.is_working = True

    flow_name = sheet["Образовательный партнёр"][row]
    education_center = EducationCenter.objects.filter(
        flow_name=flow_name
    )
    if len(education_center) == 0:
        application.delete()
        return {"status": "EdCenterMissing", "value": flow_name}
    application.education_center = education_center[0]
    contract = sheet["Номер договора на организацию обучения"][row]
    if contract != "" and contract is not None and application.education_center != None:
        contract = create_contract(
            contract, application.education_center
        )
        application.contract = contract
    program_flow_id = sheet["Идентификатор образовательной программы"][row]
    application.education_program = get_education_program(program_flow_id)
    if application.education_program is None:
        application.delete()
        education_center_year = EducationCenterProjectYear.objects.get(
            ed_center=education_center[0], project_year=project_year
        )
        if education_center_year.is_federal:
            citizen.delete()
            return {"status": "EdProgramMissingFed", "value": program_flow_id}
        return {"status": "EdProgramMissing", "value": program_flow_id}
    price = sheet["Стоимость договора гражданина"][row]
    price = 0 if price == "" or price is None else price.split(".")[0]
    application.price = int(price)
    application.save()
    return {"status": True, "value": application, "is_new": is_new}

def create_group(sheet, row, application):
    flow_id = sheet["Идентификатор потока"][row]
    group, is_new = Group.objects.get_or_create(flow_id=flow_id)
    
    if sheet["Начало обучения"][row] != None:
        group.start_date = datetime.strptime(sheet["Начало обучения"][row], "%d.%m.%Y")
    if sheet["Окончание обучения"][row] != None:
        group.end_date = datetime.strptime(sheet["Окончание обучения"][row], "%d.%m.%Y")
    group.education_program = application.education_program
    group.students.add(application)
    group.save()
    return {"status": True, "value": group, "is_new": is_new}

def create_contract(number: str, ed_center: EducationCenter):
    project_year = ProjectYear.objects.get(year=2023)
    ed_center = EducationCenterProjectYear.objects.get(
        project_year=project_year,
        ed_center=ed_center
    )
    contract, is_new = Contract.objects.get_or_create(
        number=number,
        ed_center=ed_center,
        project_year=project_year
    )
    return contract

    
def get_education_program(program_flow_id):
    education_program = EducationProgram.objects.filter(
        flow_id=program_flow_id)
    return education_program[0] if len(education_program) != 0 else None