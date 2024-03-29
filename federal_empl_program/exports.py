from datetime import datetime
import os
from zipfile import ZipFile

from django.db.models import Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils.encoding import escape_uri_path
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

from education_centers.models import EducationProgram


def quota_request(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Центры обучения"
    col_titles = [
        "Наименование Центра обучения",
        "Количество академических часов",
        "Стоимость программы",
        "Запрашиваемая квота",
        "Сумма, руб."
    ]
    
    for col_number, col_title in enumerate(col_titles, start=1):
        ws.cell(row=1, column=col_number, value=col_title)

    row_number = 2
    for center_request in request.centers_requests.all():
        for center_request in request.centers_requests.all():
            ed_center = center_request.ed_center_year.ed_center
            for program_request in center_request.quota_requests.all().order_by('program__duration'):
                quota_price = round((program_request.price / 0.93*100),0) / 100
                quota_amount = program_request.ro_quota
                quota_sum = quota_price * quota_amount
                if quota_sum != 0:
                    ws.cell(row=row_number, column=1, value=ed_center.name)
                    ws.cell(row=row_number, column=2, value=program_request.program.duration)
                    ws.cell(row=row_number, column=3, value=quota_price)
                    ws.cell(row=row_number, column=4, value=quota_amount)
                    ws.cell(row=row_number, column=5, value=quota_sum)
                    row_number += 1
    
    wb.template = False
    response = HttpResponse(
        content=save_virtual_workbook(wb), 
        content_type='application/vnd.openxmlformats-\
        officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = "attachment; filename=" + \
        escape_uri_path(f'kvota_{request.send_date}.xlsx')
    return response


def get_archive(path, name):
    zip_file = open(path, 'rb')
    response = HttpResponse(content=zip_file, content_type='application/force-download; charset=utf-8')
    time_now = datetime.now().strftime("%d/%m/%y %H:%M:%S")
    response['Content-Disposition'] = "attachment; filename=" + \
        escape_uri_path(f'{name} ({time_now}).zip')
    return response


def net_agreements_archives(agreements, type='agreements'):
    path_to_archive = "media/archive/network_agreements.zip"
    if type == 'agreements':
        files = get_agreements_files(agreements)
        archive_name = "Сетевые соглашения"
    elif type == 'programs':
        files = get_pko_programs_files(agreements)
        archive_name = "Программы ПКО"
    elif type == 'irpo_programs':
        files = get_irpo_programs_files(agreements)
        archive_name = "Программы (ИРПО) + рецензии"

    with ZipFile(path_to_archive, 'w') as archive:
        for file in files:
            archive.write(*file)

    return get_archive(path_to_archive, archive_name)


def get_irpo_programs_files(agreements):
    irpo_programs_files = []
    for agreement in agreements:
        ed_center = agreement.ed_center_year.ed_center.short_name
        programs = agreement.programs.all()
        for program in programs:
            if len(program.program_name) > 200:
                path = f'6320046206 {program.get_program_type_display()} «{program.program_name[:197]}...» {program.duration}ч'
            else:
                path = f'6320046206 {program.get_program_type_display()} «{program.program_name}» {program.duration}ч'
            program_files = []
            if program.program_word not in ['', None]:
                file_name, file_extension = os.path.splitext(program.program_word.name)
                program_files.append([program.program_word.name, f'{path}/{path}{file_extension}'])
            if program.program_pdf not in ['', None]:
                file_name, file_extension = os.path.splitext(program.program_pdf.name)
                program_files.append([program.program_pdf.name, f'{path}/{path}{file_extension}'])
            if program.teacher_review not in ['', None]:
                file_name, file_extension = os.path.splitext(program.teacher_review.name)
                program_files.append([program.teacher_review.name, f'{path}/{path} (рецензия преподавателя){file_extension}'])
            if program.employer_review not in ['', None]:
                file_name, file_extension = os.path.splitext(program.teacher_review.name)
                program_files.append([program.employer_review.name, f'{path}/{path} (рецензия работодателя){file_extension}'])

            for program in program_files:
                irpo_programs_files.append(program)
    
    return irpo_programs_files


def get_pko_programs_files(agreements):
    argeements_files = []
    for agreement in agreements:
        ed_center = agreement.ed_center_year.ed_center.short_name
        programs = agreement.programs.exclude(Q(program_file=None) | Q(program_file=''))
        for program in programs:
            file_name, file_extension = os.path.splitext(program.program_file.name)
            file_name = f'Программа {program.get_program_type_display()} «{program.program_name}» ({program.duration})'
            destination = f'{ed_center}/{file_name}{file_extension}'
            argeements_files.append([program.program_file.name, destination])
    
    return argeements_files


def get_agreements_files(agreements):
    agreements_w_file = agreements.exclude(agreement_file='')
    argeements_files = []
    for agreement in agreements_w_file:
        file_name, file_extension = os.path.splitext(agreement.agreement_file.name)
        if agreement.suffix is None:
            number = f'{agreement.agreement_number}СЗ'
        else:
            number = f'{agreement.agreement_number}СЗ{agreement.suffix}'
        destination = f'сетевое_соглашение_№{number}{file_extension}'
        argeements_files.append([agreement.agreement_file.name, destination])

    return argeements_files


def programs(agreements):
    wb = Workbook()
    ws = wb.active
    ws.title = "Учебные программы"
    col_titles = [
        "№ п/п",
        "Наименование программы",
        "ЦО",
        "Наименование профессии",
        "Вид программы (подвид для ДПО)",
        "Количество часов",
        "Форма обучения",
        "Сетевая форма реализации (да/нет)",
        "Номер договора о сетевом взаимодействии",
        "Субъект(ы) РФ, в которых планируется реализация образовательной программы"
    ]
    
    for col_number, col_title in enumerate(col_titles, start=1):
        ws.cell(row=1, column=col_number, value=col_title)
        ws.cell(row=2, column=col_number, value=col_number)

    row_number = 3
    programs = EducationProgram.objects.filter(new_agreements__in=agreements)
    for program in programs:
        if program.program_type == 'DPOPK':
            program_type = "Дополнительное профессиональное образование (повышение квалификации)"
        elif program.program_type == 'DPOPP':
            program_type = "Дополнительное профессиональное образование (профессиональная переподготовка)"
        elif program.program_type == 'POPP':
            program_type = "Профессиональное обучение (переподготовка)"
        elif program.program_type == 'POP':
            program_type = "Профессиональное обучение (профессиональная подготовка)" 
        elif program.program_type == 'POPK':
            program_type ="Профессиональное обучение (повышение квалификации)"
        agreement = program.new_agreements.first()
        if agreement.suffix == None:
            number = f'{agreement.agreement_number}/СЗ' 
        else: number = f'{agreement.agreement_number}/СЗ{agreement.suffix}'
        ws.cell(row=row_number, column=1, value=row_number - 2)
        ws.cell(row=row_number, column=2, value=program.program_name)
        ws.cell(row=row_number, column=3, value=program.ed_center.short_name)
        ws.cell(row=row_number, column=4, value=program.profession)
        ws.cell(row=row_number, column=5, value=program_type)
        ws.cell(row=row_number, column=6, value=program.duration)
        ws.cell(row=row_number, column=7, value=program.get_education_form_display())
        ws.cell(row=row_number, column=8, value="да")
        ws.cell(row=row_number, column=9, value=number)
        ws.cell(row=row_number, column=10, value="Самарская область")
        row_number += 1

    wb.template = False
    response = HttpResponse(
        content=save_virtual_workbook(wb), 
        content_type='application/vnd.openxmlformats-\
        officedocument.spreadsheetml.sheet'
    )
    time_now = datetime.now().strftime("%d/%m/%y %H:%M:%S")
    response['Content-Disposition'] = "attachment; filename=" + \
        escape_uri_path(f'programs_list ({time_now}).xlsx')
    return response


def get_full_program_type(program_type):
    if program_type == 'DPOPK':
        return "Дополнительное профессиональное образование (повышение квалификации)"
    elif program_type == 'DPOPP':
        return "Дополнительное профессиональное образование (профессиональная переподготовка)"
    elif program_type == 'POPP':
       return "Профессиональное обучение (переподготовка)"
    elif program_type == 'POP':
        return "Профессиональное обучение (профессиональная подготовка)" 
    return "Профессиональное обучение (повышение квалификации)"


def programs_w_people(agreements):
    programs = EducationProgram.objects.filter(new_agreements__in=agreements)
    wb = Workbook()
    ws = wb.active
    ws.title = "Учебные программы"
    col_titles = [
        "Фамилия, имя, отчество",
        "Уровень образования (ВО или СПО), специальность и квалификация по диплому",
        "Наличие опыта педагогической и/или производственной деятельности (указать стаж и должность)",
        "Сетевая форма организации сотрудничества (да/нет)",
        "Номер договора о сетевом взаимодействии"
    ]
    for col_number, col_title in enumerate(col_titles, start=1):
        ws.cell(row=1, column=col_number, value=col_title)

    row_number = 2
    for row_program, program in enumerate(programs, start=1):
        ws.cell(row=row_number, column=1, value=f'{row_program}. {get_full_program_type(program.program_type)} «{program.program_name}»')
        ws.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=5)
        row_number += 1
        agreement = program.new_agreements.first()
        if agreement.suffix == None:
            number = f'{agreement.agreement_number}/СЗ' 
        else: number = f'{agreement.agreement_number}/СЗ{agreement.suffix}'
        for row_teacher, teacher in enumerate(program.teachers.exclude(consent_file=None), start=1):
            ws.cell(row=row_number, column=1, value=f'{row_teacher}. {teacher.get_name()}')
            ws.cell(row=row_number, column=2, value=f'{teacher.get_education_level_display()}\n{teacher.education_major}')
            ws.cell(row=row_number, column=3, value=f'{teacher.position}\n{teacher.experience}')
            ws.cell(row=row_number, column=4, value='да')
            ws.cell(row=row_number, column=5, value=number)
            row_number += 1
    
    wb.template = False
    response = HttpResponse(
        content=save_virtual_workbook(wb), 
        content_type='application/vnd.openxmlformats-\
        officedocument.spreadsheetml.sheet'
    )
    time_now = datetime.now().strftime("%d/%m/%y %H:%M:%S")
    response['Content-Disposition'] = "attachment; filename=" + \
        escape_uri_path(f'Список кадров ({time_now}).xlsx')
    return response

def programs_w_workshops(agreements):
    programs = EducationProgram.objects.filter(new_agreements__in=agreements)
    wb = Workbook()
    ws = wb.active
    ws.title = "Учебные программы"
    col_titles = [
        "1",
        "Наименование вида образования, профессия, подвид дополнительного образования",
        "Наименование объекта, подтверждающего наличие материально-технического обеспечения, с перечнем основного оборудования",
        "Реквизиты заключения Государственной инспекции безопасности дорожного движения Министерства внутренних дел Российской Федерации о соответствии учебно-материальной базы установленным требованиям (при наличии образовательных программ подготовки водителей автомототранспортных средств)",
        "Сетевая форма реализации (да/нет)",
        "Номер договора о сетевом взаимодействии"
    ]
    for col_number, col_title in enumerate(col_titles, start=1):
        ws.cell(row=1, column=col_number, value=col_title)
        ws.cell(row=2, column=col_number, value=col_number)

    row_number = 3
    for row_program, program in enumerate(programs, start=1):
        agreement = program.new_agreements.first()
        if agreement.suffix == None:
            number = f'{agreement.agreement_number}/СЗ' 
        else: number = f'{agreement.agreement_number}/СЗ{agreement.suffix}'
        name = f'{get_full_program_type(program.program_type)} «{program.program_name}»'
        if program.program_type not in ['DPOPK', 'DPOPP']:
            name += f'({program.profession})'
        for workshop in program.workshops.exclude(address=None).exclude(name=None).exclude(programs=None).exclude(address=""):
            ws.cell(row=row_number, column=1, value=f'{row_number - 2}.')
            ws.cell(row=row_number, column=2, value=name)
            ws.cell(row=row_number, column=3, value=f'{workshop.address}\n{workshop.name}\n{workshop.equipment}')
            ws.cell(row=row_number, column=5, value="да")
            ws.cell(row=row_number, column=6, value=number)
            row_number += 1
    
    wb.template = False
    response = HttpResponse(
        content=save_virtual_workbook(wb), 
        content_type='application/vnd.openxmlformats-\
        officedocument.spreadsheetml.sheet; charset=utf-8'
    )
    time_now = datetime.now().strftime("%d/%m/%y %H:%M:%S")
    response['Content-Disposition'] = "attachment; filename=" + \
        escape_uri_path(f'список МТО ({time_now}).xlsx')
    return response