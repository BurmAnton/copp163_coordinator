import base64
import math
import os
from io import BytesIO
from django.shortcuts import get_object_or_404

import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook, load_workbook
from docx import Document as Document_compose
from docxcompose.composer import Composer
from docxtpl import DocxTemplate

from education_centers.models import DocumentType
from federal_empl_program.models import ActivityCompetence, ActivityCompetenceIndicators, ActivityType, EducationCenterProjectYear, FgosStandart, NetworkAgreement, Profstandart, ProgramDocumentation, ProgramModule, ProjectYear, Subject


def get_graph():
    buffer = BytesIO()
    plt.savefig(buffer, format='png')
    buffer.seek(0)
    image_png = buffer.getvalue()
    graph = base64.b64encode(image_png)
    graph = graph.decode('utf-8')
    buffer.close()
    return graph

def addlabels(x,y):
    for i in range(len(x)):
        plt.text(i,y[i]-45,y[i], ha = 'center', color='white', fontsize=12, fontweight=500)

def get_applications_plot(month, applications):
    plt.switch_backend('AGG')

    x = np.arange(len(month)) # the label locations
    width = 0.45 # the width of the bars
    multiplier = 0.5

    fig, ax = plt.subplots(layout='constrained')
    for attribute, measurement in applications.items():
        offset = width * multiplier
        if attribute == 'Подали заявку':
            color = '#426cf8'
        else:
            color = '#394959'
        rects = ax.bar(x + offset, measurement, width, label=attribute, color=color)
        ax.bar_label(rects, padding=2)
        multiplier += 1

    fig.set_figwidth(10)
    fig.set_figheight(5)
    fig.set_facecolor('#e9ecf7')
    ax.set_facecolor('#e9ecf7')
    ax.spines[['right', 'top']].set_visible(False)
    bar_labels = month
    

    ax.bar(applications, month, label=bar_labels, color=bar_colors)
    addlabels(applications, month)
    ax.set_ylabel('Количество заявок')
    ax.set_title('')

    ax.set_ylabel('Заявки')
    ax.set_xticks(x + width, month)
    ax.legend(loc='upper left', ncols=2)
    plt.tight_layout()
    graph = get_graph()
    return graph

def get_flow_applications_plot(weeks, weeks_stat):
    plt.switch_backend('AGG')

    x = np.arange(len(weeks))
    width = 0.2 # the width of the bars
    multiplier = 0

    fig, ax = plt.subplots(layout='constrained')
    for attribute, measurement in weeks_stat.items():
        offset = width * multiplier
        if attribute == 'Начали обучение':
            color = '#426cf8'
        elif attribute == 'Завершили обучение':
            color = '#02509e'
        else:
            color = '#394959'
        rects = ax.bar(x + offset, measurement, width, label=attribute, color=color)
        ax.bar_label(rects, padding=2)
        multiplier += 1
    
    fig.set_figwidth(12)
    fig.set_figheight(5)
    fig.set_facecolor('#F2F2F2')
    ax.set_facecolor('#F2F2F2')
    ax.spines[['right', 'top']].set_visible(False)
    bar_labels = weeks

    ax.set_ylabel('Количество заявок')
    ax.set_title('')

    ax.set_ylabel('Заявки')
    ax.set_xticks(x + width, weeks)
    ax.legend(loc='upper left', ncols=4)
    ax.legend(ncols=4, bbox_to_anchor=(0, 1),
              loc='lower left')
    plt.tight_layout()
    graph = get_graph()
    return graph

def count_levels(directory_path):
    levels_values = []
    for filename in os.listdir(directory_path):
        f = os.path.join(directory_path, filename)
        if '.xlsx' not in filename:
            print(filename)
        elif os.path.isfile(f):
            workbook = load_workbook(f)
            sheet = workbook.active
            sheet_values = []
            sheet_values.append(sheet.cell(row = 5, column = 3).value)
            sheet_values.append(sheet.cell(row = 7, column = 3).value)
            sheet_values.append(sheet.cell(row = 8, column = 3).value)
            sheet_values.append(sheet.cell(row = 9, column = 3).value)
            sheet_values.append(filename)
            levels_values.append(sheet_values)
            
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Сводная уровней школ"
    ws.cell(row=1, column=1, value="Название школы")
    ws.cell(row=1, column=2, value="Базовый")
    ws.cell(row=1, column=3, value="Основной")
    ws.cell(row=1, column=4, value="Продвинутый")
    ws.cell(row=1, column=5, value="Название файла")
    for row_number, sheet in enumerate(levels_values, start=2):
        for col_number, value in enumerate(sheet, start=1):
            ws.cell(row=row_number, column=col_number, value=value)
    wb.save('Сводная уровней школ.xlsx')


def save_program_stage(form, program):
    stage = form['save-stage']
    if stage == "1":

        program.duration_days = math.ceil(program.duration / 8)

        program.standart = FgosStandart.objects.get(id=form['standart'])
        program.profstandart = Profstandart.objects.get(id=form['profstandart'])
        program.qual_level = form['qual_level']
        program.assigned_qualif = form['assigned_qualif']
        program.gen_functions = form['gen_functions']
        program.current_control = form['current_control']
        program.middle_control = form['middle_control']
        program.final_control = form['final_control']
        program.final_control_matereils = form['final_control_matereils']
        program.final_control_criteria = form['final_control_criteria']
        program.min_score = form['min_score']

        if program.status == "0" or program.status == "1":
            program.status = "2"
            stage = 2
        else:
            stage = 1
    elif stage == "2":
        if program.status == "2":
            program.status = "3"
            stage = 3
        else:
            stage = 2
    elif stage == "3":
        if program.status == "3":
            program.status = "4"
            stage = 4
        else:
            stage = 3
    elif stage == "4" or stage == "5":
        program.status = "6"
        stage = 6

    program.save()
    return int(stage)
        

def create_irpo_program(program):
    ed_center = program.ed_center
    project_year = get_object_or_404(ProjectYear, year=2023)
    center_project_year = EducationCenterProjectYear.objects.get_or_create(
                project_year=project_year, ed_center=ed_center)[0]
    agreement = NetworkAgreement.objects.get(
        ed_center_year=center_project_year
    )
    if agreement.suffix is None or agreement.suffix == "":
        net_number = agreement.agreement_number
    else:
        net_number = f'{agreement.agreement_number}{agreement.suffix}'

    context = {
        "program": program,
        "duration_days": program.duration_days,
        "net_number": net_number,
        "schedule_dict": generate_calendar_schedule(program.duration, program)

    
    }
    if program.program_type == "DPOPK":
        doc_type = get_object_or_404(DocumentType, name="Программа ДПО ПК (ИРПО)")
    elif program.program_type == "DPOPP":   
        doc_type = get_object_or_404(DocumentType, name="Программа ДПО ПП (ИРПО)")
    else:
        doc_type = get_object_or_404(DocumentType, name="Программа ПО (ИРПО)")

    document = DocxTemplate(doc_type.template)
    document.render(context)

    path = f'media/programs/gen_irpo/{ed_center.id}'
    isExist = os.path.exists(path)
    if not isExist:
        os.makedirs(path)

    if len(program.name) > 100:
        document_name = f'6320046206 {program.get_program_type_display()} {program.name[:97]}... {program.duration}'
    else:
        document_name = f'6320046206 {program.get_program_type_display()} {program.name} {program.duration}'
    document_name = document_name.replace(',', '')
    path_to_contract = f'{path}/{document_name}.docx'
    document.save(path_to_contract)
    
    return document, document_name


def generate_calendar_schedule(duration, program):
    days = math.ceil(duration / 8)

    day_totals = [0] * days
    schedule_dict = {}
    day = 0
    schedule_dict['subjects'] = {}
    for module in program.modules.all():
        schedule_dict[module.id] = {'module':[0] * days}
        for subject in module.subjects.all():
            schedule_dict['subjects'][subject.id] = [0] * days
            subj_duration = subject.get_full_duration()
            if day_totals[day] + subj_duration < 8:
                day_totals[day] += subj_duration
                schedule_dict[module.id]['module'][day] += subj_duration
                schedule_dict['subjects'][subject.id][day] += subj_duration
            elif day_totals[day] + subj_duration == 8:
                day_totals[day] += subj_duration
                schedule_dict[module.id]['module'][day] += subj_duration
                schedule_dict['subjects'][subject.id][day] += subj_duration
                day += 1
            else: 
                while subj_duration != 0:
                    if subj_duration + day_totals[day] > 8:
                        day_subj = 8 - day_totals[day]
                    else: day_subj = subj_duration

                    day_totals[day] += day_subj
                    schedule_dict[module.id]['module'][day] += day_subj
                    schedule_dict['subjects'][subject.id][day] += day_subj
                    subj_duration -= day_subj
                    if day_totals[day] == 8: day += 1

        if module.attest_form is not None:
            schedule_dict[module.id]['attest'] = [0] * days
            test_duration = module.get_int_ex_duration()
            if day_totals[day] + test_duration < 8:
                day_totals[day] += test_duration
                schedule_dict[module.id]['module'][day] += test_duration
                schedule_dict[module.id]['attest'][day] += test_duration
            elif day_totals[day] + subj_duration == 8:
                day_totals[day] += subj_duration
                schedule_dict[module.id]['module'][day] += test_duration
                schedule_dict[module.id]['attest'][day] += test_duration
                day += 1
            else: 
                while test_duration != 0:
                    if test_duration + day_totals[day] > 8:
                        day_test = 8 - day_totals[day]
                    else: day_test = test_duration

                    day_totals[day] += day_test
                    schedule_dict[module.id]['module'][day] += day_test
                    schedule_dict[module.id]['attest'][day] += day_test
                    test_duration -= day_test
                    
                    if day_totals[day] == 8: day += 1
    schedule_dict['attest'] = [0] * days
    f_test_duration = program.get_full_ex_duration()
    if day_totals[day] + f_test_duration < 8:
        day_totals[day] += f_test_duration
        schedule_dict['attest'][day] += f_test_duration
    elif day_totals[day] + f_test_duration == 8:
        day_totals[day] += f_test_duration
        schedule_dict['attest'][day] += f_test_duration
    else: 
        while f_test_duration != 0:
                if f_test_duration + day_totals[day] > 8:
                    day_test = 8 - day_totals[day]
                else: day_test = f_test_duration

                day_totals[day] += day_test
                schedule_dict['attest'][day] += day_test
                f_test_duration -= day_test
                if day_totals[day] == 8: day += 1

    schedule_dict['total'] = day_totals
    return schedule_dict

                    
def copy_program_content(program_to, program_from):
    program_to.assigned_qualif = program_from.assigned_qualif
    program_to.duration_days = math.ceil(program_to.duration / 8)
    program_to.status = "6"
    program_to.gen_functions = program_from.gen_functions
    program_to.current_control = program_from.middle_control
    program_to.final_control = program_from.final_control
    program_to.final_control_matereils = program_from.final_control_matereils
    program_to.final_control_criteria = program_from.final_control_criteria
    program_to.qual_level = program_from.qual_level
    program_to.profstandart = program_from.profstandart
    program_to.standart = program_from.standart
    program_to.authors.add(*program_from.authors.all())
    program_to.exam_lections_duration = program_from.exam_lections_duration
    program_to.exam_practice_duration = program_from.exam_practice_duration
    program_to.exam_consultations_duration = program_from.exam_consultations_duration
    program_to.exam_independent_duration = program_from.exam_independent_duration
    program_to.exam_attest_form = program_from.exam_attest_form
    program_to.min_score = program_from.min_score
    program_to.activities.all().delete()
    program_to.documentation.all().delete()
    program_to.modules.all().delete()
    program_to.save()

    indicators = []
    
    for activity in program_from.activities.all():
        n_activity = ActivityType.objects.get(id=activity.id)
        n_activity.program = program_to
        n_activity.pk = None
        n_activity.save()
        for competence in activity.competencies.all():
            n_competence = ActivityCompetence.objects.get(id=competence.id)
            n_competence.activity = n_activity
            n_competence.pk = None
            n_competence.save()
            n_competence.equipments.add(*competence.equipments.all())
            n_competence.save()
            for indicator in competence.indicators.all():
                n_indicator = indicator
                n_indicator.competence = n_competence
                n_indicator.pk = None
                indicators.append(n_indicator)
    ActivityCompetenceIndicators.objects.bulk_create(indicators)

    docs = []
    for doc in program_from.documentation.all():
        n_doc = doc
        n_doc.program = program_to
        n_doc.pk = None
        docs.append(n_doc)
    ProgramDocumentation.objects.bulk_create(docs)

    for module in program_from.modules.all():
        n_module = ProgramModule.objects.get(id=module.id)
        n_module.pk = None
        n_module.program = program_to
        n_module.save()
        for subject in module.subjects.all():
            n_subject = subject
            n_subject.pk = None
            n_subject.module = n_module
            n_subject.save()