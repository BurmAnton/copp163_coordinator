from openpyxl import load_workbook
import django.contrib.auth.models

from citizens.models import Citizen
from federal_empl_program.models import Application, Group, Questionnaire


def import_in_db_gd(form):
    data = form.cleaned_data
    workbook = load_workbook(form.cleaned_data['import_file'])
    sheet = workbook.active
    row_count = sheet.max_row
    for row in range(2, row_count+1):
        if sheet[f"L{row}"].value != None:
            snils_number = sheet[f"L{row}"].value
            citizen = Citizen.objects.filter(snils_number=snils_number)
            if len(citizen) != 0:
                citizen = citizen[0]
                date = sheet[f"AS{row}"].value
                if date != None:
                    application = Application.objects.filter(applicant=citizen, creation_date=date)
                    if len(application) != 0:
                        application=application[0]
                        application.legacy_id = int(sheet[f"A{row}"].value)
                        if sheet[f"D{row}"].value != None:
                            purpose = get_goal(sheet, row)
                            questionnaire = Questionnaire(
                                applicant=application,
                                purpose=purpose
                            )
                            questionnaire.save()
                        if sheet[f"E{row}"].value:
                            citizen.copp_registration = True
                        if sheet[f"F{row}"].value:
                            citizen.social_status = 'UEMP'
                        if application.group is not None:
                            ed_type = sheet[f"BD{row}"].value
                            group = Group.objects.get(students=application)
                            if ed_type == 'Да':
                                group.distance_education = True
                                group.save()
                            elif ed_type == 'Смешанно':
                                group.mixed_education = True
                                group.save()
                            if sheet[f"BG{row}"].value != None:
                                citizen.education_type = get_education(sheet, row)
                            if sheet[f"BI{row}"].value:
                                application.is_working = True
                            if sheet[f"BJ{row}"].value:
                                application.find_work = 'GAJ'
                            sp_group = django.contrib.auth.models.Group.objects.filter(name='Специалист по работе с клиентами')
                            if sheet[f"BK{row}"].value != None and len(sp_group) !=0:
                                name = sheet[f"BK{row}"].value.split()
                                managers = django.contrib.auth.models.User.objects.filter(groups=sp_group[0])
                                for manager in managers:
                                    if manager.first_name == name[1] and manager.last_name == name[0]:
                                        application.citizen_consultant = manager
                            if sheet[f"BN{row}"].value:
                                application.ib_course = True
                            if sheet[f"BP{row}"].value:
                                application.pasport = True
                            #if sheet[f"BT{row}"].value:
                                #Application.consent_pers_data = True
                            if sheet[f"BV{row}"].value:
                                application.education_document = True
                            if sheet[f"BX{row}"].value:
                                application.is_enrolled = True
                            if sheet[f"BZ{row}"].value :
                                application.is_deducted = True
                            if sheet[f"BR{row}"].value == 'Трёхсторонний':
                                if application.is_working:
                                    application.contract_type = 'OLD'
                                else:
                                    application.contract_type = 'NEW'
                            elif sheet[f"BR{row}"].value == 'Двусторонний':
                                application.contract_type = 'SELF'
                        gd_status = sheet[f"C{row}"].value
                        if gd_status != None:
                            application = set_application_status_gd(gd_status, application)
                        citizen.save()
                        application.save()
                    citizen.save()
                        

def get_goal(sheet, row):
    goals_variants = [
        ('RECA', "самозанятый"),
        ('CONT', "сохранить работу"),
        ('RECD', "найти работу"),
        ('CONF', "просто поучиться"),
        ('CONF', "повысить квалификацию"),
    ]
    goal = sheet[f"D{row}"].value
    for variant in goals_variants:
        if variant[1] == goal:
            return variant[0]
    return ""

def get_education(sheet, row):
    education = sheet[f"BG{row}"].value
    education_variants = Citizen.EDUCATION_CHOICES
    for variant in education_variants:
        if variant[1] == education.capitalize():
            return variant[0]
    return ""

def set_application_status_gd(gd_status, application):
    if gd_status == "заявка получена":
        admit_status = 'RECA'
        appl_status = 'NEW'
    elif gd_status == "связались":
        admit_status = 'CONT'
        appl_status = 'VER'
    elif gd_status == "прислал часть документов":
        admit_status = 'RECD'
        appl_status = 'VER'
    elif gd_status == "допущен":
        admit_status = 'ADM'
        appl_status = 'ADM'
    elif gd_status == "начал обучение":
        admit_status = 'ADM'
        appl_status = 'SED'
    elif gd_status == "завершил обучение":
        admit_status = 'CONT'
        appl_status = 'COMP'
    elif gd_status == "отчислен":
        admit_status = application.admit_status
        appl_status = 'NCOM'
    elif gd_status == "не допущен":
        admit_status = application.admit_status
        appl_status = 'NADM'
    elif gd_status == "резерв":
        admit_status = application.admit_status
        appl_status = 'RES'
    elif gd_status == "дубликат":
        admit_status = application.admit_status
        appl_status = 'DUPL'
    elif gd_status == "другой ФО":
        admit_status = application.admit_status
        appl_status = 'OTH'
    application.appl_status = appl_status
    application.admit_status = admit_status
    application.save()
    return application
