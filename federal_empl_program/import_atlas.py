from datetime import datetime
import math
from openpyxl import load_workbook

from citizens.models import Citizen
from education_centers.models import EducationProgram, Group
from federal_empl_program.models import ApplStatus, Application, CitizenCategory, ProjectYear


def get_sheet(form):
    workbook = load_workbook(form.cleaned_data['import_file'])
    sheet = workbook.active
    return sheet


def cheak_col_match(sheet, fields_names_set):
    i = 0
    col_count = sheet.max_column
    sheet_fields = []
    sheet_col = {}
    if sheet[f"A2"].value is None:
        return {'status': 'Error', 'error_type': 'EmptySheet'}
    try:
        for col_header in range(1, col_count+1):
            if sheet.cell(row=1,column=col_header).value is not None:
                sheet_fields.append(sheet.cell(row=1,column=col_header).value)
                sheet_col[col_header] = sheet.cell(row=1,column=col_header).value
        missing_fields = []
        for field in fields_names_set:
            if field not in sheet_fields:
                missing_fields.append(field)
        if len(missing_fields) != 0:
            return {
                'status': 'Error',
                'error_type': 'MissingFieldsError',
                'missing_fields': missing_fields
            } 
    except IndexError:
        return {'status': 'Error','error_type': 'IndexError'} 
    return {'status': 'Success','sheet_col': sheet_col} 


def load_worksheet_dict(sheet, fields_names_set):
    row_count = sheet.max_row
    sheet_dict = {}
    for col in fields_names_set:
        sheet_dict[fields_names_set[col]] = []
        for row in range(2, row_count+1): 
            snils = sheet[f"A{row}"].value
            if snils != None:
                cell_value = sheet.cell(row=row,column=col).value
                try: cell_value = str(math.floor(cell_value))
                except (ValueError, TypeError): pass
                sheet_dict[fields_names_set[col]].append(cell_value)
    return sheet_dict


def import_atlas(form):
    try:
        sheet = get_sheet(form)
    except IndexError:
        return ['Import', 'IndexError']
    
    #Требуемые поля таблицы
    fields_names = {
        "Фамилия",	"Имя", "Отчество", "СНИЛС", "Пол", "Регион",
        "Email", "Контактная информация (телефон)",
        "Статус заявки в Атлас", "Статус заявки в РР",
        "Начало периода обучения", "Окончание периода обучения",
        "Программа обучения", "Категория гражданина", 
        "Дата подачи заявки на РР", "Номер заявления на РР", "ID программы в заявке"
    }

    match_result = cheak_col_match(sheet, fields_names)
    if match_result['status'] == 'Error':
        return match_result
    
    sheet = load_worksheet_dict(sheet, match_result['sheet_col'])

    applications = {'added': set(), 'updated': set()}
    citizens = {'added': set(), 'updated': set()}
    project_year = ProjectYear.objects.get(year=2024)
    project_year.appls_last_update = datetime.now()
    project_year.save()
    programs_404 = set()
    for row in range(len(sheet['Номер заявления на РР'])):
        program = get_program(sheet['ID программы в заявке'][row])
        if program is None:
            programs_404.add(f'{sheet['Программа обучения'][row]} ({sheet['ID программы в заявке'][row]})')
        else:
            citizen, is_new = get_citizen(sheet, row)
            if is_new: citizens['added'].add(citizen)
            else: citizens['updated'].add(citizen)
            application, is_new = Application.objects.get_or_create(
                project_year=project_year,
                applicant=citizen,
            )
            if is_new: 
                update_application(application, sheet, row)
                applications['added'].add(citizen)
            elif (sheet['Статус заявки в Атлас'][row] != 'Отклонена'
            or sheet["Статус заявки в РР"][row] != 'Услуга прекращена'):
                update_application(application, sheet, row)
                applications['updated'].add(citizen)
            elif (application.rvr_status == 'Услуга прекращена' 
            or application.atlas_status == 'Отклонена') \
            or application.atlas_id == sheet["Номер заявления на РР"][row]: 
                update_application(application, sheet, row)
                applications['updated'].add(citizen)
        
    return {
        'status': 'OK', 
        'citizens': citizens, 
        'applications': applications,
        'programs_404': programs_404
    }


def update_application(appl, sheet, row):   
    appl.atlas_id = sheet["Номер заявления на РР"][row]
    appl.rvr_status = sheet["Статус заявки в РР"][row]
    appl.atlas_status = sheet["Статус заявки в Атлас"][row]
    appl.status = get_appl_status(appl.rvr_status, appl.atlas_status)
    appl.education_program = get_program(sheet["ID программы в заявке"][row])
    appl.group = get_group(sheet, row, appl.education_program)
    appl.education_center = appl.group.education_center
    appl.citizen_category = get_category(sheet["Категория гражданина"][row])
    appl.save()

    return appl


def extract_date(date):
    return datetime.strptime(date, "%d.%m.%Y")


def get_group(sheet, row, program):
    start_date = extract_date(sheet["Начало периода обучения"][row])
    end_date = extract_date(sheet["Окончание периода обучения"][row])
    name = f'{program.program_name} (с {start_date.strftime('%d/%m')} по {start_date.strftime('%d/%m')})'
    group, _ = Group.objects.get_or_create(
        name=name,
        education_program=program,
        is_atlas=True, 
        start_date=start_date, 
        end_date=end_date
        )
    
    return group


def get_category(category_name):
    project_year = ProjectYear.objects.get(year=2024)
    category, _ = CitizenCategory.objects.get_or_create(official_name=category_name, project_year=project_year)
    return category


def get_program(atlas_id):
    try:
        return EducationProgram.objects.get(atlas_id=atlas_id, is_atlas=True)
    except:
        return None


def get_citizen(sheet, row):
    sex = sheet["Пол"][row]
    if sex == "Мужской": sex = 'M'
    else: sex = 'F'
    citizen, is_new = Citizen.objects.get_or_create(
        snils_number = sheet["СНИЛС"][row],
    )
    citizen.first_name = sheet["Имя"][row].title()
    citizen.last_name = sheet["Фамилия"][row].title()
    citizen.middle_name = sheet["Отчество"][row].title()
    citizen.sex = sex
    citizen.email = sheet["Email"][row].lower()
    citizen.phone_number = sheet["Контактная информация (телефон)"][row]
    citizen.res_region =  sheet["Регион"][row]
    citizen.save()
    return [citizen, is_new]
    

new_s = [
    "Новая", "Требуется личная явка", "Принято в работу", 
    "Требуется проведение видеоконференцсвязи", "Приглашение отправлено", "Ожидание загрузки документов",
]
apprv_s = [
    "Договор ожидает подписания", "Договор на подписании", 
    "Одобрено центром занятости населения", "Ожидает подписи соглашения", 
    "Ожидание загрузки заявления", "Документы на иправлении"
]
empl_contract_s = [
    "Заключён договор", "Ожидание загрузки договора", "Ожидает проверки договора", "Ожидает проверки заявления",
    "Заявление на иправлении", "Договор на иправлении"
]
ed_contract_s = ["Ожидает начала обучения", "Ожидает приказ на зачисление", "Ожидает прикрепления приказа на зачисление"]
study_s = ["Обучается", "Проходит обучение"] #"Проходит обучение"
cancel_s = ["Услуга прекращена", "Отклонена", "Отчислен"]
end_s = ['Услуга оказана', 'Запрос оригиналов документа о квалификации', 'Ожидает документа о квалификации']
  
def get_appl_status(rvr_status, atlas_status):
    short_name = "undefined"
    if rvr_status in end_s:
        short_name = "end"
    elif rvr_status in cancel_s or atlas_status in cancel_s:
        short_name = "canceled"
    elif rvr_status in study_s or atlas_status in study_s:
        short_name = "studying"
    elif atlas_status in ed_contract_s or rvr_status in ed_contract_s:
        short_name = "ed_contract"
    elif rvr_status in empl_contract_s or atlas_status in empl_contract_s:
        short_name = "empl_contract"
    elif rvr_status in apprv_s or rvr_status in apprv_s:
        short_name = "approved"
    elif rvr_status in new_s or atlas_status in new_s:
        short_name = "new"

    return ApplStatus.objects.get(short_name=short_name)