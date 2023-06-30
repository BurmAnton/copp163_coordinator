import math
from openpyxl import load_workbook

from .models import Competence, EducationProgram, EducationCenter


def get_sheet(form):
    workbook = load_workbook(form.cleaned_data['import_file'])
    sheet = workbook.active
    return sheet

def cheak_col_match(sheet, fields_names_set):
    i = 0
    col_count = sheet.max_column
    sheet_fields = []
    sheet_col = {}
    if sheet[f"A2"].value is None:
        return ['Import', 'EmptySheet']
    try:
        for col_header in range(1, col_count+1):
            if sheet.cell(row=1,column=col_header).value is not None:
                sheet_fields.append(sheet.cell(row=1,column=col_header).value)
                sheet_col[col_header] = sheet.cell(row=1,column=col_header).value
        missing_fields = []
        for field in fields_names_set:
            if field not in sheet_fields:
                missing_fields.append(field)
        if len(missing_fields) != 0:
            return ['Import', 'MissingFieldsError', missing_fields]
    except IndexError:
            return ['Import', 'IndexError']
    return [True, sheet_col]

def load_worksheet_dict(sheet, fields_names_set):
    row_count = sheet.max_row
    sheet_dict = {}
    for col in fields_names_set:
        sheet_dict[fields_names_set[col]] = []
        for row in range(2, row_count+1): 
            snils = sheet[f"A{row}"].value
            if snils != None:
                cell_value = sheet.cell(row=row,column=col).value
                try: cell_value = str(math.floor(cell_value))
                except (ValueError, TypeError): pass
                sheet_dict[fields_names_set[col]].append(cell_value)
    return sheet_dict

def programs(form):
    try:
        sheet = get_sheet(form)
    except IndexError:
        return ['Import', 'IndexError']

    #Требуемые поля таблицы
    fields_names = {
        'Название образовательной организации', 'Направление программы', 
        'Название программы', 'Профессия','Описание компетенции', 
        'Вид программы', 'Количество часов', 'Форма обучения', 
        'Входное требование', 'Примечание'
    }

    cheak_col_names = cheak_col_match(sheet, fields_names)
    if cheak_col_names[0] != True:
        return cheak_col_names

    sheet_dict = load_worksheet_dict(sheet, cheak_col_names[1])
    
    missing_fields = []
    change_programs_count = 0
    added_programs_count = 0
    for row in range(len(sheet_dict['Название программы'])):
        program = load_program(sheet_dict, row)
        if program[0] == 'OK':
            if program[1]: added_programs_count += 1
            else: change_programs_count += 1
        elif program[0] == 'MissingField':
            missing_fields.append(program)
    return [
        'Import', 'OK', added_programs_count, 
        change_programs_count,  missing_fields
    ]

def load_program(sheet, row):
    missing_fields = []
    program_name = sheet["Название программы"][row]
    if program_name != None: 
        program_name = program_name.strip()
    ed_center_name = sheet["Название образовательной организации"][row].strip()
    if ed_center_name == "": 
        missing_fields.append("Название образовательной организации")
    ed_center = EducationCenter.objects.get_or_create(
        name=ed_center_name
    )
    competence = sheet["Направление программы"][row].strip()
    if competence == "":
        missing_fields.append("Направление программы")
    else:
        competence = Competence.objects.get_or_create(
            title=competence,
            #is_irpo=True
        )
    profession = sheet["Профессия"][row].strip()
    if profession == "":
        missing_fields.append("Профессия")
    description = sheet["Описание компетенции"][row].strip()
    if description == "":
        missing_fields.append("Описание компетенции")
    program_type = None
    if sheet["Вид программы"][row].strip() is not None:
        for program_choice in EducationProgram.PROGRAM_TYPES:
            if sheet["Вид программы"][row].strip() in program_choice[1]: 
                program_type = program_choice[0]
                break
    if program_type == None: missing_fields.append("Вид программы")
    duration = int(sheet["Количество часов"][row].strip())
    if duration == "":
        missing_fields.append("Количество часов")
    education_form = None
    if sheet["Форма обучения"][row].strip() is not None:
        for form_choice in EducationProgram.EDUCATION_FORMS:
            if sheet["Форма обучения"][row].strip() in form_choice[1]: 
                education_form = form_choice[0]
                break
    if education_form == None: missing_fields.append("Форма обучения")
    entry_requirements = None
    if sheet["Входное требование"][row].strip() is not None:
        for form_choice in EducationProgram.EDUCATION_CHOICES:
            if sheet["Входное требование"][row].strip() in form_choice[1]: 
                entry_requirements = form_choice[0]
                break
    if entry_requirements == None: missing_fields.append("Входное требование")
    notes = sheet["Примечание"][row].strip()

    if len(missing_fields) == 0:
        program, is_new = EducationProgram.objects.get_or_create(
            ed_center=ed_center,
            program_name=program_name,
            competence=competence,
            duration=duration,
            program_type=program_type
        )
        program.profession = profession
        program.description = description
        program.education_form = education_form
        program.entry_requirements = entry_requirements
        program.save()
        return ['OK', is_new, program]
    return ['MissingField', missing_fields, program_name]

def merge_ed_centers(form):
    try:
        sheet = get_sheet(form)
    except IndexError:
        return ['Import', 'IndexError']
    
    #Требуемые поля таблицы
    fields_names = {"ID", "Слияние"}

    cheak_col_names = cheak_col_match(sheet, fields_names)
    if cheak_col_names[0] != True:
        return cheak_col_names
    
    sheet_dict = load_worksheet_dict(sheet, cheak_col_names[1])

    merged_centers_count = 0
    errors = []
    for row in range(len(sheet_dict['ID'])):
        merge = merge_center(sheet_dict, row)
        if merge[0] == 'OK':
            merged_centers_count += 1
        else: errors.append(merge)
    return ['OK', merged_centers_count, errors]

def merge_center(sheet, row):
    merge_center_id = sheet["ID"][row]
    merge_ed_center = EducationCenter.objects.filter(id=merge_center_id)
    if len(merge_ed_center) == 0:
        return ['MergeIdNotFound', merge_center_id, row]
    merge_ed_center = merge_ed_center[0]

    center_id = sheet["Слияние"][row]
    ed_center = EducationCenter.objects.filter(id=center_id)
    if len(ed_center) == 0:
        return ['IdNotFound', ed_center, row]
    ed_center = ed_center[0]
    
    ed_center.teachers.add(*merge_ed_center.teachers.all())
    ed_center.workshops.add(*merge_ed_center.workshops.all())
    ed_center.programs.add(*merge_ed_center.programs.all())
    ed_center.edcenter_applicants.add(*merge_ed_center.edcenter_applicants.all())
    ed_center.ticket_programs.add(*merge_ed_center.ticket_programs.all())
    ed_center.save()
    merge_ed_center.delete()
    
    return ['OK', ed_center]