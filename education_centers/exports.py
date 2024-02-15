from datetime import datetime

from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils.encoding import escape_uri_path
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

from federal_empl_program.models import EducationCenterProjectYear, ProjectYear

from .models import EducationCenter, EducationProgram, Workshop


def ed_centers():
    ed_centers = EducationCenter.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Центры обучения"
    col_titles = [
        "ID",
        "Наименование",
        "Краткое наименование"
    ]

    for col_number, col_title in enumerate(col_titles, start=1):
        ws.cell(row=1, column=col_number, value=col_title)
    
    row_number = 2
    for ed_center in ed_centers:
        ws.cell(row=row_number, column=1, value=ed_center.id)
        ws.cell(row=row_number, column=2, value=ed_center.name)
        ws.cell(row=row_number, column=3, value=ed_center.short_name)
        row_number += 1
    
    wb.template = False
    response = HttpResponse(
        content=save_virtual_workbook(wb), 
        content_type='application/vnd.openxmlformats-\
        officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = "attachment; filename=" + \
        escape_uri_path(f'ed_centers_{datetime.now().strftime("%d/%m/%y")}.xlsx')
    return response

def programs(project_years=None, ed_centers=None):
    programs = EducationProgram.objects.all()
    project_year = get_object_or_404(ProjectYear, year=2023)
    if ed_centers != None:
        centers_project_year = EducationCenterProjectYear.objects.filter(
            project_year=project_year,
            ed_center__in=ed_centers
        )
    else: centers_project_year = EducationCenterProjectYear.objects.filter(
            project_year=project_year,
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Программы"
    col_titles = [
        "№ п/п", 
        "Центр обучения", 
        "Субъект РФ",
        "Федеральный образовательный центр",
        "Направление программы", 
        "Название программы", 
        "Профессия", 
        "Описание", 
        "Вид программы", 
        "Колво часов", 
        "Форма обучения", 
        "Входные требования",
        "Примечания"
    ]
    for col_number, col_title in enumerate(col_titles, start=1):
        ws.cell(row=1, column=col_number, value=col_title)
        ws.cell(row=2, column=col_number, value=col_number)

    row_number = 3
    for center in centers_project_year:
        if center.is_federal: is_federal = "Да"
        else: is_federal = "Нет"
        for program in center.ed_center.programs.all():
            if program.program_type == 'DPOPK':
                program_type = "Дополнительное профессиональное образование (повышение квалификации)"
            elif program.program_type == 'DPOPP':
                program_type = "Дополнительное профессиональное образование (профессиональная переподготовка)"
            elif program.program_type == 'POPP':
                program_type = "Профессиональное обучение (переподготовка)"
            elif program.program_type == 'POP':
                program_type = "Профессиональное обучение (профессиональная подготовка)" 
            elif program.program_type == 'POPK':
                program_type ="Профессиональное обучение (повышение квалификации)"
            ws.cell(row=row_number, column=1, value=row_number-2)
            ws.cell(row=row_number, column=2, value=program.ed_center.name)
            ws.cell(row=row_number, column=3, value="Самарская область")
            ws.cell(row=row_number, column=4, value=is_federal)
            ws.cell(row=row_number, column=5, value=program.competence.title)
            ws.cell(row=row_number, column=6, value=program.program_name)
            ws.cell(row=row_number, column=7, value=program.profession)
            ws.cell(row=row_number, column=8, value=program.description)
            ws.cell(row=row_number, column=9, value=program_type)
            ws.cell(row=row_number, column=10, value=program.duration)
            ws.cell(row=row_number, column=11, 
                    value=program.get_education_form_display())
            ws.cell(row=row_number, column=12, 
                    value=program.get_entry_requirements_display())
            ws.cell(row=row_number, column=13, value=program.notes)
            row_number += 1
    
    wb.template = False
    response = HttpResponse(
        content=save_virtual_workbook(wb), 
        content_type='application/vnd.openxmlformats-\
        officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = "attachment; filename=" + \
        escape_uri_path(f'Programs_{datetime.now().strftime("%d/%m/%y")}.xlsx')
    return response

def workshops(project_years=2023):
    workshops = Workshop.objects.exclude(address=None).exclude(name=None).exclude(programs=None).exclude(address="")
    wb = Workbook()
    ws = wb.active
    ws.title = "Аудитории"
    col_titles = [
        "Название", 
        "Центр обучения", 
        "Адрес",
        "Программы",
        "Вид занятий",
        "Оборудование"
    ]
    for col_number, col_title in enumerate(col_titles, start=1):
        ws.cell(row=1, column=col_number, value=col_title)
    
    row_number = 2
    for workshop in workshops:
        ws.cell(row=row_number, column=1, value=workshop.name)
        ws.cell(row=row_number, column=2, value=workshop.education_center.name)
        ws.cell(row=row_number, column=3, value=workshop.address)
        programs = workshop.programs.all().values_list("program_name", flat=True)
        programs = '\n'.join(list(programs))
        ws.cell(row=row_number, column=4, value=programs)
        ws.cell(row=row_number, column=5, value=workshop.get_classes_type_display())
        ws.cell(row=row_number, column=6, value=workshop.equipment)
        row_number += 1

    wb.template = False
    response = HttpResponse(
        content=save_virtual_workbook(wb), 
        content_type='application/vnd.openxmlformats-\
        officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = "attachment; filename=" + \
        escape_uri_path(f'Workshops ({datetime.now().strftime("%d/%m/%y %H:%M:%S")}).xlsx')
    return response
