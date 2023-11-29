
from datetime import date, datetime

from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils.encoding import escape_uri_path
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

from future_ticket.models import SchoolProjectYear, TicketProjectYear

from .models import School


def schools():
    wb = Workbook()
    ws = wb.active
    ws.title = "Школы"
    col_titles = [
        "Наименование",
        "ИНН",
        "Тер. управление",
        "Отправили заявку",
    ]
    
    for col_number, col_title in enumerate(col_titles, start=1):
        ws.cell(row=1, column=col_number, value=col_title)

    row_number = 2
    schools = School.objects.all()
    for school in schools:
        application = 'Нет'
        project_year = get_object_or_404(TicketProjectYear, year=2023)
        school_year = SchoolProjectYear.objects.filter(
            school=school,
            project_year=project_year
        )
        if len(school_year) != 0: application = 'Да'
        ws.cell(row=row_number, column=1, value=school.name)
        ws.cell(row=row_number, column=2, value=school.inn)
        ws.cell(row=row_number, column=3, 
                value=school.get_territorial_administration_display())
        ws.cell(row=row_number, column=4, value=application)
        row_number += 1
    
    wb.template = False
    response = HttpResponse(
        content=save_virtual_workbook(wb), 
        content_type='application/vnd.openxmlformats-\
        officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = "attachment; filename=" + \
        escape_uri_path(f'schools_list_{date.today()}.xlsx')
    return response