import math

from django.db import IntegrityError
from django.shortcuts import get_object_or_404
from openpyxl import load_workbook

from citizens.models import Municipality, School


def get_sheet(form):
    workbook = load_workbook(form.cleaned_data['import_file'])
    sheet = workbook.active
    return sheet

def cheak_col_match(sheet, fields_names_set):
    i = 0
    col_count = sheet.max_column
    sheet_fields = []
    sheet_col = {}
    if sheet[f"A2"].value is None:
        return ['Import', 'EmptySheet']
    try:
        for col_header in range(1, col_count+1):
            if sheet.cell(row=1,column=col_header).value is not None:
                sheet_fields.append(sheet.cell(row=1,column=col_header).value)
                sheet_col[col_header] = sheet.cell(row=1,column=col_header).value
        missing_fields = []
        for field in fields_names_set:
            if field not in sheet_fields:
                missing_fields.append(field)
        if len(missing_fields) != 0:
            return ['Import', 'MissingFieldsError', missing_fields]
    except IndexError:
            return ['Import', 'IndexError']
    return [True, sheet_col]

def load_worksheet_dict(sheet, fields_names_set):
    row_count = sheet.max_row
    sheet_dict = {}
    for col in fields_names_set:
        sheet_dict[fields_names_set[col]] = []
        for row in range(2, row_count+1): 
            snils = sheet[f"A{row}"].value
            if snils != None:
                cell_value = sheet.cell(row=row,column=col).value
                try: cell_value = str(math.floor(cell_value))
                except (ValueError, TypeError): pass
                sheet_dict[fields_names_set[col]].append(cell_value)
    return sheet_dict

def schools(form):
    try:
        sheet = get_sheet(form)
    except IndexError:
        return ['Import', 'IndexError']

    #Требуемые поля таблицы
    fields_names = {"ИНН", "ТУ/ДО", "Муниципалитет", "Название ОО"}

    cheak_col_names = cheak_col_match(sheet, fields_names)
    if cheak_col_names[0] != True:
        return cheak_col_names

    sheet_dict = load_worksheet_dict(sheet, cheak_col_names[1])

    missing_fields = []
    added_schools_count = 0
    duplicates = []
    for row in range(len(sheet_dict['ИНН'])):
        school = load_school(sheet_dict, row)
        if school[0] == 'OK':
            if school[1]:
                added_schools_count += 1
        elif school[0] == 'MissingField':
            missing_fields.append(school)
        elif school[0] == 'Duplicate':
            duplicates.append(school)
    return ['OK', added_schools_count, missing_fields, duplicates]

def load_school(sheet, row):
    missing_fields = []

    inn = sheet["ИНН"][row]
    if inn != "" and inn != None: 
        inn = inn.strip()
    else: missing_fields.append("ИНН")
    name = sheet["Название ОО"][row]
    if name != "" and name != None: 
        name = name.strip()
    else: missing_fields.append("Название ОО")
    territorial_administration = None
    for form_choice in School.TER_CHOICES:
        if sheet['ТУ/ДО'][row].strip() in form_choice[1]: 
            territorial_administration = form_choice[0]
            break
    if territorial_administration == None: missing_fields.append("ТУ/ДО")
    municipality_name= sheet["Муниципалитет"][row]
    if municipality_name != "" and municipality_name != None: 
        municipality_name = municipality_name.strip()
    else: missing_fields.append("Муниципалитет")

    if len(missing_fields) == 0:
        municipality, is_new = Municipality.objects.get_or_create(
            name=municipality_name
        )
        try:
            school, is_new = School.objects.get_or_create(
                inn=inn,
                name=name,
                territorial_administration=territorial_administration,
                municipality=municipality
            )
        except IntegrityError:
            return ['Duplicate', inn, row]
        return ['OK', is_new, school]
    return ['MissingField', missing_fields, row]