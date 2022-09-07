from openpyxl import load_workbook

from .models import PartnerContact, PartnerOrganization, PartnerContactEmail, PartnerContactPhone, Project

def get_sheet(form):
    workbook = load_workbook(form.cleaned_data['import_file'])
    sheet = workbook.active
    return sheet

def cheak_col_match(sheet, fields_names_set):
    i = 0
    col_count = sheet.max_column
    sheet_fields = []
    sheet_col = {}
    if sheet[f"A2"].value is None:
        return [False, 'EmptySheet']
    try:
        for col_header in range(1, col_count+1):
            if sheet.cell(row=1,column=col_header).value is not None:
                sheet_fields.append(sheet.cell(row=1,column=col_header).value)
                sheet_col[col_header] = sheet.cell(row=1,column=col_header).value
        missing_fields = []
        for field in fields_names_set:
            if field not in sheet_fields:
                missing_fields.append(field)
        if len(missing_fields) != 0:
            return [False, 'FieldError', missing_fields]
    except IndexError:
            return [False, 'IndexError']
    return [True, sheet_col]

def load_worksheet_dict(sheet, fields_names_set):
    row_count = sheet.max_row
    sheet_dict = {}
    for col in fields_names_set:
        sheet_dict[fields_names_set[col]] = []
        for row in range(2, row_count+1): 
            school_name = sheet[f"A{row}"].value
            if school_name != None:
                sheet_dict[fields_names_set[col]].append(sheet.cell(row=row,column=col).value)
    return sheet_dict

def contacts(form):
    try:
        sheet = get_sheet(form)
    except IndexError:
        return [False, 'IndexError']
    
    #Требуемые поля таблицы
    fields_names = {
        'Фамилия', 'Имя', 'Отчество', 
        'Должность', 'Название организации', 'ИНН организации', 
        'Телефоны', 'Emails', 'Проекты', 'Комментарий', 
    }

    cheak_col_names = cheak_col_match(sheet, fields_names)
    if cheak_col_names[0] != True:
        return cheak_col_names

    sheet_dict = load_worksheet_dict(sheet, cheak_col_names[1])
    contacts_count = 0
    new_organizations = 0
    problems = []
    for row in range(len(sheet_dict['Фамилия'])):
        #Загрузка организации
        organization = load_organization(sheet_dict, row)
        if organization[0] == "NewOrganization":
            new_organizations += 1
        organization = organization[1]
        #Загрузка контакта
        contact = load_contact(organization, sheet_dict, row)
        if contact[0] == "OK":
            contacts_count += 1
        else:
            problems.append(contact)
    return [True, contacts_count, problems, new_organizations]

def load_organization(sheet, row):
    try: organization_inn = str(int(sheet["ИНН организации"][row]))
    except: organization_inn = str(sheet["ИНН организации"][row])
    organization_inn = organization_inn.replace(" ", "")
    cheak_organization = PartnerOrganization.objects.filter(organization_inn=organization_inn)
    if len(cheak_organization) > 0:
        organization = cheak_organization[0]
        return ["OldOrganization", organization]
    name = sheet["Название организации"][row].strip()
    organization_type = sheet["Тип организации"][row].strip()
    ORG_TYPES = {
        'ЦО СПО': 'ECSPO',
        'ЦО частные': 'ECP',
        'Гос. орган': 'GOV',
        'СОУ': 'SCHL',
        'Другие': 'OTH'
    }
    organization = PartnerOrganization(
        name=name,
        organization_inn=organization_inn,
        organization_type=ORG_TYPES[organization_type]
    )
    organization.save()
    return ["NewOrganization", organization]

def load_contact(organization, sheet, row):
    #Проверяем есть ли такой контакт
    last_name = sheet["Фамилия"][row].replace(" ", "").capitalize()
    first_name = sheet["Имя"][row].replace(" ", "").capitalize()
    middle_name = sheet["Отчество"][row].replace(" ", "").capitalize()
    cheak_contact = PartnerContact.objects.filter(
        last_name=last_name, 
        first_name=first_name,
        middle_name=middle_name,
        organization=organization
    )
    if len(cheak_contact) == 0:
        #Находим проекты из списка
        projects_list = []
        projects = sheet["Проекты"][row].strip().split(',')
        for project_name in projects:
            project_name = project_name.strip()
            project = Project.objects.filter(project_name=project_name)
            if len(project) == 0:
                return ['ProjectError', project_name]
            projects_list.append(project[0])
        #Добавляем контакт
        job_title = sheet["Должность"][row].strip()
        commentary = sheet["Комментарий"][row].strip()
        contact = PartnerContact(
            last_name=last_name, 
            first_name=first_name,
            middle_name=middle_name,
            organization=organization,
            job_title=job_title,
            commentary=commentary
        )
        contact.save()
        contact.projects.add(*projects_list)
        #Добавляем номера из списка
        phones = sheet["Телефоны"][row].strip().split(',')
        for phone in phones:
            phone = phone.strip()
            phone = PartnerContactPhone(contact=contact, phone=phone)
            phone.save()
        #Добавляем почты из списка
        emails = sheet["Emails"][row].strip().split(',')
        for email in emails:
            email = email.strip()
            email = PartnerContactEmail(contact=contact, email=email)
            email.save()
        return ['OK', contact]
    return ['ContactDublicateError', cheak_contact[0]]