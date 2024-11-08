from datetime import datetime

from django.db.models import Sum
from django.http import HttpResponse
from django.utils.encoding import escape_uri_path
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

from .models import TicketEvent, TicketProfession, TicketProgram, TicketQuota


def professions():
    professions = TicketProfession.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Професии"
    col_titles = [
        "ID", 
        "Название", 
        "Среда",
        "Федеральная?",
        "Колво программ",
        "Колво заявок (Запрос)",
        "Колво заявок (Одобрено)"
    ]
    for col_number, col_title in enumerate(col_titles, start=1):
        ws.cell(row=1, column=col_number, value=col_title)

    row_number = 2
    for profession in professions:
        if profession.is_federal: is_federal = "Да"
        else: is_federal = "Нет"
        quota_sum = TicketQuota.objects.filter(profession=profession).distinct(
                            ).aggregate(Sum('value'), Sum('approved_value'))
        value = quota_sum['value__sum']
        approved_value = quota_sum['approved_value__sum']
        if value == None: value = 0
        if approved_value == None: approved_value = 0
        ws.cell(row=row_number, column=1, value=profession.id)
        ws.cell(row=row_number, column=2, value=profession.name)
        ws.cell(row=row_number, column=3, value=profession.prof_enviroment.name)
        ws.cell(row=row_number, column=4, value=is_federal)
        ws.cell(row=row_number, column=5, value=len(profession.programs.all()))
        ws.cell(row=row_number, column=6, value=value)
        ws.cell(row=row_number, column=7, value=approved_value)
        row_number += 1
    
    wb.template = False
    response = HttpResponse(
        content=save_virtual_workbook(wb), 
        content_type='application/vnd.openxmlformats-\
        officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = "attachment; filename=" + \
        escape_uri_path(
            f'professions_{datetime.now().strftime("%d/%m/%y %H:%M")}.xlsx'
        )
    return response

def programs():
    programs = TicketProgram.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Программы"
    col_titles = [
        "ЦО", 
        "Профессия",
        "Среда",
        "Статус",
        "Краткое описание задания",
        "Ссылка на программу",
        "Возрастные категории",
        "Автор",
        "Телефон",
        "Email"
    ]
    for col_number, col_title in enumerate(col_titles, start=1):
        ws.cell(row=1, column=col_number, value=col_title)

    row_number = 2
    for program in programs:
        age_groups = ""
        for index, age_group in enumerate(program.age_groups.all()):
            age_groups += f'{age_group.name}, '
            if index == len(program.age_groups.all()) - 1:
                age_groups += age_group.name
        ws.cell(row=row_number, column=1, value=program.ed_center.short_name)
        ws.cell(row=row_number, column=2, value=program.profession.name)
        ws.cell(row=row_number, column=3, value=program.profession.prof_enviroment.name)
        ws.cell(row=row_number, column=4, value=program.get_status_display())
        ws.cell(row=row_number, column=5, value=program.description)
        ws.cell(row=row_number, column=6, value=program.program_link)
        ws.cell(row=row_number, column=7, value=age_groups)
        ws.cell(row=row_number, column=8, value=program.author.teacher.get_name())
        ws.cell(row=row_number, column=9, value=program.author.phone)
        ws.cell(row=row_number, column=10, value=program.author.email)
        row_number += 1
    
    wb.template = False
    response = HttpResponse(
        content=save_virtual_workbook(wb), 
        content_type='application/vnd.openxmlformats-\
        officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = "attachment; filename=" + \
        escape_uri_path(
            f'programs_{datetime.now().strftime("%d/%m/%y %H:%M")}.xlsx'
        )
    return response

def schools_applications(applications):
    wb = Workbook()
    ws = wb.active
    ws.title = "Школы"
    col_titles = [
        "Школа", 
        "ИНН",
        "Тер. управление",
        "ФИО",
        "Должность",
        "Email",
        "Телефон",
        "Приказ"
    ]

    for col_number, col_title in enumerate(col_titles, start=1):
        ws.cell(row=1, column=col_number, value=col_title)

    row_number = 2
    for application in applications:
        ws.cell(row=row_number, column=1, value=application.school.name)
        ws.cell(row=row_number, column=2, value=application.school.inn)
        ws.cell(
            row=row_number, 
            column=3, 
            value=application.school.get_territorial_administration_display()
        )
        ws.cell(row=row_number, column=4, value=application.resp_full_name)
        ws.cell(row=row_number, column=5, value=application.resp_position)
        ws.cell(row=row_number, column=6, value=application.email)
        ws.cell(row=row_number, column=7, value=application.phone)
        ws.cell(
            row=row_number, 
            column=8, 
            value=f'https://copp63-coordinator.ru{application.resp_order.url}'
        )
        row_number += 1
    
    wb.template = False
    response = HttpResponse(
        content=save_virtual_workbook(wb), 
        content_type='application/vnd.openxmlformats-\
        officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = "attachment; filename=" + \
        escape_uri_path(
            f'prof_min_{datetime.now().strftime("%d/%m/%y %H:%M")}.xlsx'
        )
    return response
        

def qoutas(quotas):
    wb = Workbook()
    ws = wb.active
    ws.title = "Программы"
    
    col_titles = [
        "ЦО", 
        "Тер. управление",
        "Запрос",
        "Распределенно",
        "Выполнено",
        "Школа",
        "Профессия"
    ]
    for col_number, col_title in enumerate(col_titles, start=1):
        ws.cell(row=1, column=col_number, value=col_title)

    for row, quota in enumerate(quotas, start=2):
        ws.cell(row=row, column=1, value=quota.ed_center.short_name)
        ws.cell(row=row, column=2, value=quota.school.get_territorial_administration_display())
        ws.cell(row=row, column=3, value=quota.value)
        ws.cell(row=row, column=4, value=quota.reserved_quota)
        ws.cell(row=row, column=5, value=quota.completed_quota)
        ws.cell(row=row, column=6, value=quota.school.name)
        ws.cell(row=row, column=7, value=quota.profession.name)

    wb.template = False
    response = HttpResponse(
        content=save_virtual_workbook(wb), 
        content_type='application/vnd.openxmlformats-\
        officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = "attachment; filename=" + \
        escape_uri_path(
            f'quotas_{datetime.now().strftime("%d/%m/%y %H:%M")}.xlsx'
        )
    return response

def events():
    wb = Workbook()
    ws = wb.active
    ws.title = "Мероприятия"
    col_titles = [
        "ЦО",
        "Профессия",
        "Дата проведения",
        "Выделено квоты",
        "Выполнено квоты",
        "Ссылка",
    ]
    for col_number, col_title in enumerate(col_titles, start=1):
        ws.cell(row=1, column=col_number, value=col_title)

    events = TicketEvent.objects.exclude(participants_limit=0)
    for row, event in enumerate(events, start=2):
        ws.cell(row=row, column=1, value=event.ed_center.ed_center.short_name)
        ws.cell(row=row, column=2, value=event.profession.name)
        ws.cell(row=row, column=3, value=event.event_date.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=4, value=event.participants_limit)
        participants = event.participants.filter(
                is_double=False, is_attend=True
            ).count()
        if participants >= event.participants_limit:
            ws.cell(row=row, column=5, value=event.participants_limit)
        else:
            ws.cell(row=row, column=5, value=participants)
        ws.cell(row=row, column=6, value=event.photo_link)
    wb.template = False
    response = HttpResponse(
        content=save_virtual_workbook(wb), 
        content_type='application/vnd.openxmlformats-\
        officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = "attachment; filename=" + \
        escape_uri_path(
            f'events_{datetime.now().strftime("%d/%m/%y %H:%M")}.xlsx'
        )
    return response