import math
from django.shortcuts import get_object_or_404
from openpyxl import load_workbook
from citizens.models import School
from education_centers.models import EducationCenter, Teacher

from future_ticket.models import AgeGroup, ProfEnviroment, ProgramAuthor, TicketProfession, TicketProgram


def get_sheet(form):
    workbook = load_workbook(form.cleaned_data['import_file'])
    sheet = workbook.active
    return sheet

def cheak_col_match(sheet, fields_names_set):
    i = 0
    col_count = sheet.max_column
    sheet_fields = []
    sheet_col = {}
    if sheet[f"A2"].value is None:
        return ['Import', 'EmptySheet']
    try:
        for col_header in range(1, col_count+1):
            if sheet.cell(row=1,column=col_header).value is not None:
                sheet_fields.append(sheet.cell(row=1,column=col_header).value)
                sheet_col[col_header] = sheet.cell(row=1,column=col_header).value
        missing_fields = []
        for field in fields_names_set:
            if field not in sheet_fields:
                missing_fields.append(field)
        if len(missing_fields) != 0:
            return ['Import', 'MissingFieldsError', missing_fields]
    except IndexError:
            return ['Import', 'IndexError']
    return [True, sheet_col]

def load_worksheet_dict(sheet, fields_names_set):
    row_count = sheet.max_row
    sheet_dict = {}
    for col in fields_names_set:
        sheet_dict[fields_names_set[col]] = []
        for row in range(2, row_count+1): 
            snils = sheet[f"A{row}"].value
            if snils != None:
                cell_value = sheet.cell(row=row,column=col).value
                try: cell_value = str(math.floor(cell_value))
                except (ValueError, TypeError): pass
                sheet_dict[fields_names_set[col]].append(cell_value)
    return sheet_dict

def change_professions(form):
    try:
        sheet = get_sheet(form)
    except IndexError:
        return ['Import', 'IndexError']
    
    #Требуемые поля таблицы
    fields_names = {
        "ID", "Слияние", "Название", "Среда", "Федеральная?", "Удалить"
    }
    
    cheak_col_names = cheak_col_match(sheet, fields_names)
    if cheak_col_names[0] != True:
        return cheak_col_names
    
    sheet_dict = load_worksheet_dict(sheet, cheak_col_names[1])

    missing_fields = []
    errors = []
    changed_professions_count = 0
    delete = 0
    for row in range(len(sheet_dict['ID'])):
        profession = change_profession(sheet_dict, row)
        if profession[0] == 'OK':
            if profession[1]:
                if profession[2] == "Delete": delete += 1
                else: changed_professions_count += 1
        elif profession[0] == 'MissingField':
            missing_fields.append(profession)
        else: errors.append(profession)
    return ['OK', changed_professions_count, missing_fields, errors, delete]

def change_profession(sheet, row):
    missing_fields = []

    profession_id = sheet["ID"][row]
    profession = TicketProfession.objects.filter(id=profession_id)

    if len(profession) != 0: 
        profession = profession[0]
    else: 
        missing_fields.append("ID")
        return ['MissingField', missing_fields, row + 2]

    if sheet["Удалить"][row] == '1':
        if len(profession.programs.all()) == 0:
            profession.delete()
            return ['OK', True, "Delete", row + 2]
        return ['ProgramsAttached', profession, row + 2]

    name = sheet["Название"][row]
    if name != "" and name != None: name = name.strip()
    else: missing_fields.append("Название")

    prof_enviroment = sheet["Среда"][row]
    prof_enviroment = ProfEnviroment.objects.filter(name=prof_enviroment)
    if len(prof_enviroment) != 0: 
        prof_enviroment = prof_enviroment[0]
    else: missing_fields.append("Среда")

    is_federal = sheet["Федеральная?"][row]
    if is_federal == "Да": is_federal = True
    elif is_federal == "Нет": is_federal = False
    else: missing_fields.append("Федеральная?")
    
    if len(missing_fields) == 0:
        profession.name = name
        profession.prof_enviroment = prof_enviroment
        profession.is_federal = is_federal
        profession.save()

        merge_id = sheet["Слияние"][row]
        if merge_id != "" and merge_id != None:
            merge_profession = TicketProfession.objects.filter(id=merge_id)
            if len(merge_profession) == 0: 
                return ['FailedMerge', profession, row]
            else:
                merge_profession = merge_profession[0]
                for program in profession.programs.all():
                    program.profession = merge_profession
                    program.save()
                for quota in profession.quotas.all():
                    quota.profession = merge_profession
                    quota.save()
                profession.delete()
                return ['OK', True, merge_profession, row + 2]
        return ['OK', False, profession]
    return ['MissingField', missing_fields, row + 2]

def professions(form):
    try:
        sheet = get_sheet(form)
    except IndexError:
        return ['Import', 'IndexError']

    #Требуемые поля таблицы
    fields_names = {"Наименование","Среда"}

    cheak_col_names = cheak_col_match(sheet, fields_names)
    if cheak_col_names[0] != True:
        return cheak_col_names

    sheet_dict = load_worksheet_dict(sheet, cheak_col_names[1])

    missing_fields = []
    errors = []
    added_professions_count = 0
    for row in range(len(sheet_dict['Наименование'])):
        profession = load_profession(sheet_dict, row)
        if profession[0] == 'OK':
            if profession[1]:
                added_professions_count += 1
        elif profession[0] == 'MissingField':
            missing_fields.append(profession)
        else: errors.append(profession)
    return ['OK', added_professions_count, missing_fields, errors]

def load_profession(sheet, row):
    missing_fields = []

    name = sheet["Наименование"][row]
    if name != "" and name != None: 
        name = name.strip().capitalize()
    else: missing_fields.append("Наименование")
    prof_enviroment_name = sheet["Среда"][row]
    if prof_enviroment_name != "" and prof_enviroment_name != None: 
        prof_enviroment_name = prof_enviroment_name.strip().lower()
    else: missing_fields.append("Среда")
    if len(missing_fields) == 0:
        prof_enviroment = ProfEnviroment.objects.filter(
                name=prof_enviroment_name
            )
        if len(prof_enviroment) == 0: 
            return ['WrongEnviroment', prof_enviroment_name, row]
        prof_enviroment = prof_enviroment.first()
        profession, is_new = TicketProfession.objects.get_or_create(
            name=name,
            prof_enviroment=prof_enviroment,
            is_federal=False
        )
        return ['OK', is_new, profession]
    return ['MissingField', missing_fields, row]

def programs(form):
    try:
        sheet = get_sheet(form)
    except IndexError:
        return ['Import', 'IndexError']

    #Требуемые поля таблицы
    fields_names = {
        'Профессиональная среда', 'Профессия',
        'Краткое описание задания', 'ФИО автора программы',
        'Должность автора программы', 'Телефон автора программы',
        'Почта автора программы', 'Место работы автора программы',
        'Формат проведения', 'Возрастная категория 6-7',
        'Возрастная категория 8-9', 'Возрастная категория 10-11',
        'Доступность для участников с ОВЗ', 'Файл программы (ссылка)'
    }

    cheak_col_names = cheak_col_match(sheet, fields_names)
    if cheak_col_names[0] != True:
        return cheak_col_names

    sheet_dict = load_worksheet_dict(sheet, cheak_col_names[1])

    missing_fields = []
    errors = []
    change_programs_count = 0
    added_programs_count = 0
    for row in range(len(sheet_dict['Профессия'])):
        program = load_program(sheet_dict, row)
        if program[0] == 'OK':
            if program[1]: added_programs_count += 1
            else: change_programs_count += 1
        elif program[0] == 'MissingField':
            missing_fields.append(program)
        else: errors.append(program)
    return [
        'Import', 'OK', added_programs_count, 
        change_programs_count,  missing_fields, errors
    ]

def load_program(sheet, row):
    missing_fields = []

    profession_name = sheet["Профессия"][row]
    if profession_name != "" and profession_name != None: 
        profession_name = profession_name.strip().capitalize()
    else: missing_fields.append("Профессия")
    prof_enviroment_name = sheet["Профессиональная среда"][row]
    if prof_enviroment_name != "" and prof_enviroment_name != None: 
        prof_enviroment_name = prof_enviroment_name.strip().lower()
    else: missing_fields.append("Профессиональная среда")
    description = sheet["Краткое описание задания"][row]
    if description == "" or description == None: 
        missing_fields.append("Краткое описание задания")  
    author = sheet['ФИО автора программы'][row]
    if author != "" and author != None and len(author.split()) > 2:
        author = author.title()
    else: missing_fields.append("ФИО автора программы")  
    position = sheet['Должность автора программы'][row]
    if position != "" and position != None:
        position = position.strip()
    else: missing_fields.append("Должность автора программы")
    phone = sheet['Телефон автора программы'][row]
    if phone != "" and phone != None:
        phone = phone.strip()
    else: missing_fields.append("Телефон автора программы")
    email = sheet['Почта автора программы'][row]
    if email != "" and email != None:
        email = email.strip()
    else: missing_fields.append("Почта автора программы")
    ed_center_name = sheet['Место работы автора программы'][row]
    if ed_center_name != "" and ed_center_name != None:
        ed_center_name = ed_center_name.strip()
    else: missing_fields.append("Место работы автора программы")
    education_form = None
    for form_choice in TicketProgram.EDUCATION_FORMS:
        if sheet['Формат проведения'][row].strip() in form_choice[1]: 
            education_form = form_choice[0]
            break
    if education_form == None: missing_fields.append("Формат проведения")
    age_group = None
    age_groups = []
    if sheet['Возрастная категория 6-7'][row] == 'Да':
        age_groups.append('6-7 класс')
    if sheet['Возрастная категория 8-9'][row] == 'Да':
        age_groups.append('8-9 класс')
    if sheet['Возрастная категория 10-11'][row] == 'Да':
        age_groups.append('10-11 класс')
    program_link = sheet['Файл программы (ссылка)'][row]

    if len(missing_fields) == 0:
        ed_center = EducationCenter.objects.filter(short_name=ed_center_name)
        if len(ed_center) == 0: return ['MissingEdCenter', ed_center_name, row]
        elif len(ed_center) > 1:
            return ['ToManyEdCenters', ed_center_name, row, len(ed_center)]
        ed_center = ed_center.first()
        profession = TicketProfession.objects.filter(name=profession_name)
        if len(profession) > 0: profession = profession.first()
        else:
            prof_enviroment = ProfEnviroment.objects.filter(
                name=prof_enviroment_name
            )
            if len(prof_enviroment) == 0: 
                return ['WrongEnviroment', prof_enviroment_name, row]
            prof_enviroment = prof_enviroment.first()
            profession = TicketProfession(
                name=profession_name,
                prof_enviroment=prof_enviroment
            )
            profession.save()
        author = author.split()
        last_name = author[0]
        first_name = author[1]
        if len(author) == 3: middle_name = author[2]
        else: middle_name = ""
        teacher, is_new = Teacher.objects.get_or_create(
            organization=ed_center,
            last_name=last_name,
            first_name=first_name,
            middle_name=middle_name,
        )
        author, is_new = ProgramAuthor.objects.get_or_create(
            teacher=teacher,
            phone=phone,
            email=email
        )
        program, is_new = TicketProgram.objects.get_or_create(
            ed_center=ed_center,
            author=author,
            profession=profession,
            status='PRWD',
            education_form=education_form,
            description=description,
        )
        age_groups = AgeGroup.objects.filter(name__in=age_groups)
        program.age_groups.set(age_groups)
        program.program_link = program_link
        program.save()

        return ['OK', is_new, program]
    return ['MissingField', missing_fields, row]


def schools_address(form):
    try:
        sheet = get_sheet(form)
    except IndexError:
        return ['Import', 'IndexError']

    #Требуемые поля таблицы
    fields_names = {
        'Нас. пункт', 'ИНН', 'Юридический адрес'
    }

    cheak_col_names = cheak_col_match(sheet, fields_names)
    if cheak_col_names[0] != True:
        return cheak_col_names

    sheet = load_worksheet_dict(sheet, cheak_col_names[1])
    
    added_adress_count = 0
    missing_fields = []
    errors = []
    for row in range(len(sheet['ИНН'])):
        valid_row = True
        inn = sheet["ИНН"][row]
        school = School.objects.filter(inn=inn)
        if len(school) == 0: 
            errors.append(f'Школы с ИНН {inn} нет в БД (строка - {row+2})')
            valid_row = False
        elif len(school) > 1:
            errors.append(f'Найдено больше одной школы с ИНН: {inn} (строка - {row+2})')
            valid_row = False
        else: school = school[0]
        adress = sheet["Юридический адрес"][row]
        if adress != "" and adress != None:
            adress = adress.strip()
        else: 
            missing_fields.append([row+2, "Юридический адрес", inn])
            valid_row = False
        city = sheet["Нас. пункт"][row]
        if city != "" and city != None:
            city = city.strip()
        else: 
            missing_fields.append([row+2, "Нас. пункт", inn])
            valid_row = False
        if valid_row:
            school.adress = adress
            school.city = city
            school.save()
            added_adress_count += 1
    
    return [added_adress_count, missing_fields, errors]