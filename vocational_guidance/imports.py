from openpyxl import load_workbook

from datetime import datetime, tzinfo
from django.utils import timezone

from citizens.models import Citizen, DisabilityType, School
from users.models import User, Group
from .models import VocGuidAssessment, VocGuidTest, TimeSlot, BiletDistribution, TestContact
from education_centers.models import EducationCenter

def get_sheet(form):
    workbook = load_workbook(form.cleaned_data['import_file'])
    sheet = workbook.active
    return sheet

def bvb_teachers(form):
    try:
        sheet = get_sheet(form)
    except IndexError:
        return [False, 'IndexError']

    fields_names_set = {
        'ТУ', 'Школа', 'Населенный пункт', 
        'Логин', 'Пароль', 'Квота', 'Фед. квота'
    }

    cheak = cheak_col_match(sheet, fields_names_set)
    if cheak[0] == False:
        return cheak
    
    sheet_dict = load_worksheet_dict(sheet, cheak[1])
    teachers = 0
    updated = 0
    for row in range(len(sheet_dict['Логин'])):
        school = load_school(sheet_dict, row)
        if school[1] == "Added":
            teachers += 1
        elif school[1] == "Updated":
            updated += 1
    return [True, teachers, updated]

def cheak_col_match(sheet, fields_names_set):
    i = 0
    col_count = sheet.max_column
    sheet_fields = []
    sheet_col = {}
    if sheet[f"A2"].value is None:
        return [False, 'EmptySheet']
    try:
        for col_header in range(1, col_count+1):
            if sheet.cell(row=1,column=col_header).value is not None:
                sheet_fields.append(sheet.cell(row=1,column=col_header).value)
                sheet_col[col_header] = sheet.cell(row=1,column=col_header).value
        missing_fields = []
        for field in fields_names_set:
            if field not in sheet_fields:
                missing_fields.append(field)
        if len(missing_fields) != 0:
            return [False, 'FieldError', missing_fields]
    except IndexError:
            return [False, 'IndexError']
    return [True, sheet_col]
    
def load_worksheet_dict(sheet, fields_names_set):
    row_count = sheet.max_row
    sheet_dict = {}
    for col in fields_names_set:
        sheet_dict[fields_names_set[col]] = []
        for row in range(2, row_count+1): 
            login = sheet[f"D{row}"].value
            if login != None:
                sheet_dict[fields_names_set[col]].append(sheet.cell(row=row,column=col).value)
    return sheet_dict

def load_school(sheet_dict, row):
    school_name = sheet_dict["Школа"][row]
    territorial_administration = sheet_dict["ТУ"][row]
    if territorial_administration != "":
        for ter_admin in School.TER_CHOICES:
            if ter_admin[1] == territorial_administration:
                territorial_administration = ter_admin[0]
                break
    city = sheet_dict["Населенный пункт"][row]

    school = School.objects.filter(name=school_name, territorial_administration=territorial_administration, city=city)
    if len(school) != 0:
        school = update_school(sheet_dict, row, school[0])
        return [school, "Updated"]

    school = School.objects.filter(name=school_name, city=city, territorial_administration=None)
    if len(school) != 0:
        school = update_school(sheet_dict, row, school[0])
        return [school, "Updated"]

    school = School.objects.filter(name=school_name, city=None, territorial_administration=None)
    if len(school) != 0:
        school = update_school(sheet_dict, row, school[0])
        return [school, "Updated"]

    school = add_school(sheet_dict, row)
    return [school, 'Added']
        

def add_school(sheet_dict, row):
    school = sheet_dict["Школа"][row]
    territorial_administration = sheet_dict["ТУ"][row]
    if territorial_administration != "":
        for ter_admin in School.TER_CHOICES:
            if ter_admin[1] == territorial_administration:
                territorial_administration = ter_admin[0]
                break
    if len(territorial_administration) > 10:
        territorial_administration = ""
    city = sheet_dict["Населенный пункт"][row]
    quota = sheet_dict["Квота"][row]
    fed_quota = sheet_dict["Фед. квота"][row]
    #Создаю школу
    school = School(
        name=school,
        territorial_administration=territorial_administration,
        city=city
    )
    school.save()
    #Задаю квоту
    bilet = BiletDistribution(
        school=school,
        test_type=fed_quota,
        quota=quota
    )
    bilet.save()

    #Проверяю существует ли учитель с таким email в системе, 
    #если нет создаю нового
    teach_email = sheet_dict["Логин"][row]
    teacher = User.objects.filter(email=teach_email)
    if len(teacher) == 0:
        teacher = add_teacher(sheet_dict, row)
    else:
        teacher = teacher[0]
    #Добавляю учителя к школе
    school.school_coordinators.add(teacher)

    return school

def update_school(sheet_dict, row, school):
    #Обновляю данные школы
    territorial_administration = sheet_dict["ТУ"][row]
    if territorial_administration != "":
        for ter_admin in School.TER_CHOICES:
            if ter_admin[1] == territorial_administration:
                territorial_administration = ter_admin[0]
                break
    city = sheet_dict["Населенный пункт"][row]
    
    school.city = city
    if territorial_administration != "" and len(territorial_administration) <= 10:
        school.territorial_administration = territorial_administration
    school.city = city
    school.save()

    #Обновляю квоту
    quota = sheet_dict["Квота"][row]
    fed_quota = sheet_dict["Фед. квота"][row]

    bilet = BiletDistribution.objects.filter(school=school)
    if len(bilet) != 0:
        bilet = bilet[0]
        bilet.test_type=fed_quota
        bilet.quota=quota
        bilet.save()
    else:
        bilet = BiletDistribution(
            school=school,
            test_type=fed_quota,
            quota=quota
        )
        bilet.save()

    #Проверяю наличие координатора в школе
    if len(school.school_coordinators.all()) == 0:
    #Проверяю существует ли учитель с таким email в системе, 
    #если нет создаю нового
        teach_email = sheet_dict["Логин"][row]
        teacher = User.objects.filter(email=teach_email)
        if len(teacher) == 0:
            teacher = add_teacher(sheet_dict, row)
        else:
            teacher = teacher[0]
        #Добавляю учителя к школе
        school.school_coordinators.add(teacher)

    return school

def add_teacher(sheet_dict, row):
    email=sheet_dict["Логин"][row]
    password=sheet_dict["Пароль"][row]

    user = User.objects.create_user(email=email, password=password)
    user.save()
    group = Group.objects.filter(name="Координатор")
    if len(group) != 0:
        user.groups.add(group[0])
    user.save()

    return user

def slots_import(form):
    try:
        sheet = get_sheet(form)
    except IndexError:
        return [False, 'IndexError']

    fields_names_set = {
        'ЦО', 'Программа', 
        'Тип', 'Возрастная категория', 
        'Описание', 'Дата', 'Время', 
        'ОВЗ', 'Проф. среда', 'Профессии',
        'ФИО', 'Email', 'Телефон'
    }

    cheak = cheak_col_match(sheet, fields_names_set)
    if cheak[0] == False:
        return cheak
    
    sheet_dict = load_worksheet_dict_slots(sheet, cheak[1])
    slots = 0
    nf_ed_centers = set()
    for row in range(len(sheet_dict['Программа'])):
        test = load_slot(sheet_dict, row)
        if test[0] == "OK":
            slots += test[2]
        else:
            nf_ed_centers.add(test[1])
    return [slots, nf_ed_centers]
    
def load_worksheet_dict_slots(sheet, fields_names_set):
    row_count = sheet.max_row
    sheet_dict = {}
    for col in fields_names_set:
        sheet_dict[fields_names_set[col]] = []
        for row in range(2, row_count+1): 
            login = sheet[f"B{row}"].value
            if login != None:
                sheet_dict[fields_names_set[col]].append(sheet.cell(row=row,column=col).value)
    return sheet_dict

def load_slot(sheet_dict, row):
    ed_center = sheet_dict["ЦО"][row]
    program = sheet_dict["Программа"][row]
    time_slots = sheet_dict["Время"][row]
    date = sheet_dict["Дата"][row]

    test = load_test(sheet_dict, row)
    if test[0] == "OK":
        test = test[1]
    else:
        return ["NOT_FOUND_ED_CNTR", test[1]]

    time_slots = time_slots.split(",")
    slots_count = 0
    for slot in time_slots:
        if "с 15:00 до 16:30" in slot:
            slot = 'MID'
        elif "с 16:30 до 18:00" in slot:
            slot = 'EVN'
        elif 'с 10:00 до 11:30' in slot:
            slot = 'MRN'
        else: 
            return ["OK", test, slots_count]
        time_slot = TimeSlot.objects.filter(test=test, date=date, slot=slot)
        if len(time_slot) == 0:
            slot = TimeSlot(
                test=test, 
                date=date, 
                slot=slot
            )
            slot.save()
            slots_count += 1
    return ["OK", test, slots_count]

def load_test(sheet_dict, row):
    name = sheet_dict["Программа"][row]
    description = sheet_dict["Описание"][row]
    disabilities = sheet_dict["ОВЗ"][row]
    education_center = sheet_dict["ЦО"][row]
    age_group = sheet_dict["Возрастная категория"][row]
    category = sheet_dict["Проф. среда"][row]
    guid_type = sheet_dict['Тип'][row]
    full_name = sheet_dict['ФИО'][row]
    email = sheet_dict['Email'][row]
    phone = sheet_dict['Телефон'][row]
    profession = sheet_dict['Профессии'][row]

    if guid_type == "парк":
        guid_type = "SPO"
    else:
        guid_type = "VO"
    
    for them in VocGuidTest.THEMES_CHOICES:
        if category == them[1]:
            thematic_env = them[0]
            break
    else:
        thematic_env = ""

    education_center = EducationCenter.objects.filter(name=education_center)

    if len(education_center) > 0:
        education_center = education_center[0]
        if age_group == "6-7 класс":
            age_group = '6-7'
        elif age_group == "8-9 класс":
            age_group = '8-9'
        elif age_group == "10-11 класс":
            age_group = '10-11'
        else:
            age_group = 'ALL'

        test = VocGuidTest.objects.filter(
            name=name, 
            education_center=education_center, 
            age_group=age_group,
            guid_type=guid_type
        )

        if len(test) == 0:
            test = VocGuidTest(
                name = name,
                education_center = education_center,
                description = description,
                age_group = age_group,
                guid_type = guid_type,
                thematic_env = thematic_env
            )
            test.save()
            contact = TestContact(
                test=test,
                full_name=full_name,
                email=email,
                phone=phone
            )
            contact.save()
        else:
            test = test[0]

        if disabilities is not None:
            disabilities = disabilities.split(",")
            for disability in disabilities:
                disability = DisabilityType.objects.filter(name=disability)
                if len(disability) > 0:
                    test.disability_types.add(disability[0])
        
        if thematic_env != "":
            test.thematic_env = thematic_env
        test.profession = profession

        test.save()
        return ["OK", test]

    return ["NOT_FOUND_ED_CNTR", sheet_dict["ЦО"][row]]

def import_external_slots(form):
    try:
        sheet = get_sheet(form)
    except IndexError:
        return [False, 'IndexError']

    fields_names_set = {
        'ЦО', 'Программа', 
        'Тип', 'Возрастная категория', 
        'Описание', 'Дата', 'Время', 
        'ОВЗ', 'Проф. среда', 'Профессии',
        'ФИО', 'Email', 'Телефон'
    }

    cheak = cheak_col_match(sheet, fields_names_set)
    if cheak[0] == False:
        return cheak
    
    sheet_dict = load_worksheet_dict_slots(sheet, cheak[1])
    slots = 0
    nf_ed_centers = set()
    for row in range(len(sheet_dict['Программа'])):
        test = load_slot(sheet_dict, row)
        if test[0] == "OK":
            slots += test[2]
        else:
            nf_ed_centers.add(test[1])
    return [slots, nf_ed_centers]

def load_worksheet_dict(sheet, fields_names_set, col_name):
    row_count = sheet.max_row
    sheet_dict = {}
    for col in fields_names_set:
        sheet_dict[fields_names_set[col]] = []
        for row in range(2, row_count+1): 
            login = sheet[f"{col_name}{row}"].value
            if login != None:
                sheet_dict[fields_names_set[col]].append(sheet.cell(row=row,column=col).value)
    return sheet_dict

def matching_bvb_students(form):
    try:
        sheet = get_sheet(form)
    except IndexError:
        return [False, 'IndexError']

    fields_names_set = {
        '№пп', 'Федеральный округ',
        'Регион', 'Муниципалитет', 'ФИО', 
        'образовательная организация', 
        'класс (без буквы)', 'почта, телефон', 'наличие ОВЗ',
        'нозология', 'скан согласия родителей на обработку перс.данных',
        'количество пройденных диагностик', 'даты пройденных диагностик',
        'даты пройденных диагностик'
    }

    cheak = cheak_col_match(sheet, fields_names_set)
    if cheak[0] == False:
        return cheak
    
    sheet_dict = load_worksheet_dict(sheet, cheak[1], 'E')
    
    missing_schools = []
    import_schools_set = set()
    find_students = 0
    for row in range(len(sheet_dict['ФИО'])):
        school = sheet_dict['образовательная организация'][row]
        import_schools_set.add(school)
        full_name = sheet_dict['ФИО'][row].split(' ')
        first_name = full_name[1]
        last_name = full_name[0]
        if len(full_name) == 3:
            middle_name=full_name[2]
        else:
            middle_name = ""
        schools = School.objects.filter(name=school)
        if len(schools) != 0:
            for school in schools:
                student = Citizen.objects.filter(
                    first_name=first_name,
                    last_name=last_name,
                    middle_name=middle_name,
                    school=school
                )
                if len(student) != 0:
                    student = student[0]
                    find_students += 1
                    student_assessments = VocGuidAssessment.objects.filter(participant=student)
                    if len(student_assessments) != 0:
                        for assessment in student_assessments:
                            assessment.bilet_platform = True
                            assessment.diagnostics_count =sheet_dict['количество пройденных диагностик'][row]
                            assessment.save()
                            
    #Проверяем наличие школы с платформы в списке, фиксируем не найденные
    for school in BiletDistribution.objects.filter(quota__gt=0):
        if school.school.name not in import_schools_set:
            missing_schools.append(school.school)
    
    return [missing_schools, find_students, "OK"]
