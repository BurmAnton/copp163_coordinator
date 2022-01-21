from datetime import timedelta, date
from django.core.checks import messages
from django.db.models import Count
from django.http.response import JsonResponse
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth import authenticate, login, logout
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse
from django.urls import reverse
from django.db import IntegrityError

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from pysendpulse.pysendpulse import PySendPulse

import string
import random
import json

from .forms import ImportDataForm
from .imports import bvb_teachers, slots_import, matching_bvb_students

from .models import TimeSlot, VocGuidTest, VocGuidGroup, VocGuidAssessment, TestContact, BiletDistribution
from users.models import User, Group
from education_centers.models import EducationCenter
from citizens.models import Citizen, DisabilityType, School, SchoolClass

# Create your views here.
@login_required(login_url='signin')
def index(request):
    school_group = Group.objects.get(name='Школьник')
    coordinator_group = Group.objects.get(name='Координатор')
    if len(User.objects.filter(groups=school_group, email=request.user.email)) != 0:
        citizen = Citizen.objects.filter(
            email=request.user.email,
        )
        if len(citizen) != 0:
            citizen = citizen[0]
        else:
            return HttpResponseRedirect(reverse('signout'))
        return HttpResponseRedirect(reverse('profile', args=(citizen.id,)))
    elif len(User.objects.filter(groups=coordinator_group, email=request.user.email)) != 0:
        user = User.objects.filter(email=request.user.email)
        school = School.objects.filter(school_coordinators=user[0].id)
        return HttpResponseRedirect(reverse("school_dash", args=(school[0].id,)))
    elif request.user.is_staff:
        return HttpResponseRedirect(reverse("bilet_dashboard"))
    return HttpResponseRedirect(reverse("index"))

def quotas_dashboard(request):
    quotas = []
    all_quota = 0
    all_spent_quota = 0
    ter_admins = []
    for ter_adm in School.TER_CHOICES:
        ter_list = [ter_adm[1], []]
        ter_quota = 0
        ter_spent_quota = 0
        schools = School.objects.filter(territorial_administration=ter_adm[0])
        for school in schools:
            quota = BiletDistribution.objects.filter(school=school).values("quota")
            if len(quota) != 0:
                participants = Citizen.objects.filter(school=school)
                nonprofit_slots = TimeSlot.objects.filter(is_nonprofit=True)
                spent_quota = len(VocGuidAssessment.objects.filter(participant__in=participants, attendance=True).exclude(slot__in=nonprofit_slots))
                ter_spent_quota += spent_quota
                all_spent_quota += spent_quota
                ter_quota += quota[0]['quota']
                all_quota += quota[0]['quota']
                difference = quota[0]['quota'] - spent_quota
                if spent_quota != 0 or quota[0]['quota'] != 0:
                    ter_list[1].append([school.name, quota[0]['quota'], spent_quota, difference])
        ter_list.append([ter_spent_quota, ter_quota, ter_quota-ter_spent_quota])
        if ter_spent_quota != 0 or ter_quota != 0:
            ter_admins.append([ter_adm[0], ter_adm[1]])
            ter_list.append(ter_adm[0])
            quotas.append(ter_list)
    ter_list = ["Без тер. управления", []]
    ter_quota = 0
    ter_spent_quota = 0
    schools = School.objects.filter(territorial_administration=None)
    for school in schools:
        quota = BiletDistribution.objects.filter(school=school).values("quota")
        if len(quota) != 0:
            participants = Citizen.objects.filter(school=school)
            nonprofit_slots = TimeSlot.objects.filter(is_nonprofit=True)
            spent_quota = len(VocGuidAssessment.objects.filter(participant__in=participants, attendance=True).exclude(slot__in=nonprofit_slots))
            ter_spent_quota += spent_quota
            all_spent_quota += spent_quota
            ter_quota += quota[0]['quota']
            all_quota += quota[0]['quota']
            difference = quota[0]['quota'] - spent_quota
            if spent_quota != 0 or quota[0]['quota'] != 0:
                ter_list[1].append([school.name, quota[0]['quota'], spent_quota, difference])
    ter_list.append([ter_spent_quota, ter_quota, ter_quota-ter_spent_quota])
    if ter_spent_quota != 0 or ter_quota != 0:
        quotas.append(ter_list)

    return render(request, "vocational_guidance/dashboard_quotas.html", {
        'quotas': quotas,
        "all": [all_quota, all_spent_quota, all_quota-all_spent_quota],
        "ter_admins": ter_admins
    })

@login_required
def students_dashboard(request):
    slots = TimeSlot.objects.order_by('test', 'slot')
    assessments = VocGuidAssessment.objects.filter(attendance=True, slot__in=slots)
    slots_lists = []

    return render(request, "vocational_guidance/dashboard_students.html", {
        'assessments': assessments,
    })

@csrf_exempt
def bvb_students_report(request):
    if request.method == 'POST':
        slots = TimeSlot.objects.order_by('test', 'slot')
        assessments = VocGuidAssessment.objects.filter(attendance=True, slot__in=slots)
        wb = Workbook()
        ws = wb.active
        ws.title = "Cтуденты"
        column_names = [
            "ID", "Субъект РФ", "Муниципальный район",
            "Тип", "Формат проведения", "ТУ",
            "ЦО", "Название","Дата (дд.мм.гггг)", 
            "Время начала", "Время конца", "Школа",
            "Фамилия", "Имя", "Отчество", "Email", "BVB", 
            "Отчётная ссылка", "Подтверждён", "Адрес", 
            "Контактное лицо на площадке", "Количество мест",
            "Описание мероприятия", "Спикеры", "Профессии", 
            "Сферы", "Интервал таймслота (мин)"
        ]
        active_column = 1
        for name in column_names:
            ws.cell(row=1, column=active_column, value=name)
            active_column += 1
        
        active_row = 2
        for assessment in assessments:
            if assessment.slot.slot == "MRN":
                start="10:00"
                end="11:30"
            elif assessment.slot.slot == "MID":
                start="15:00"
                end="16:30"
            else:
                start="16:30"
                end="18:00"

            if assessment.slot.report_link == None:
                report_link="–"
            else:
                report_link=assessment.slot.report_link

            if len(TestContact.objects.filter(test=assessment.test)) != 0:
                contact = assessment.test.contact.full_name
            else:
                contact = "–"

            cell_values = {
                "ID": assessment.slot.id,
                "Субъект РФ": "Самарская область", 
                "Муниципальный район": "",
                "Тип": "Онлайн", 
                "Формат проведения": "Проба", 
                "ТУ": assessment.participant.school.get_territorial_administration_display(),
                "ЦО": assessment.test.education_center.name, 
                "Название": assessment.test.name,
                "Дата (дд.мм.гггг)": assessment.slot.date, 
                "Время начала": start, 
                "Время конца": end, 
                "Школа": f"{assessment.participant.school.name} ({assessment.participant.school.city})",
                "Фамилия": assessment.participant.last_name, 
                "Имя": assessment.participant.first_name, 
                "Отчество": assessment.participant.middle_name,
                "Email": assessment.participant.email,
                "BVB": assessment.bilet_platform, 
                "Отчётная ссылка": report_link, 
                "Подтверждён": assessment.is_checked, 
                "Адрес": "Онлайн", 
                "Контактное лицо на площадке": contact, 
                "Количество мест": 8,
                "Описание мероприятия": assessment.test.description, 
                "Спикеры": contact, 
                "Профессии": assessment.test.profession, 
                "Сферы": assessment.test.get_thematic_env_display(), 
                "Интервал таймслота (мин)": 90
            }
            active_col = 1
            for key, value in cell_values.items():
                ws.cell(row=active_row, column=active_col, value=value)
                active_col += 1
            active_row += 1

        wb.template = False
        wb.save('Report_BVB.xlsx')
        response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Report_BVB.xlsx'
        return response
    return HttpResponseRedirect(reverse('admin:index'))

def bilet_dashboard(request):
    assessments = VocGuidAssessment.objects.filter(attendance=True)
    slots_unique = set()
    for assessment in assessments:
        slots_unique.add(TimeSlot.objects.get(assessments=assessment))
    slots_count = TimeSlot.objects.filter(assessments__in=assessments)
    slots_count_link = TimeSlot.objects.filter(assessments__in=assessments).exclude(report_link=None)
    count_assessments = len(assessments)
    unique_assessments = assessments.values('participant').distinct()

    tests = VocGuidTest.objects.all()
    ed_centers = EducationCenter.objects.filter(voc_guid_sessions__in=tests)
    #Подсчёт квоты по каждому центру и пробе
    tests_dict = {}
    for center in ed_centers:
        tests_dict[center.name] = {}
        tests_dict[center.name]['count'] = 0
    for test in tests:
        tests_dict[test.education_center.name][f"{test.name} ({test.id})"] = [0, 0, 0, 0]
    for assessment in assessments:
        tests_dict[assessment.test.education_center.name][f"{assessment.test.name} ({assessment.test.id})"][0] += 1
        if assessment.slot.slot == 'MRN':
            tests_dict[assessment.test.education_center.name][f"{assessment.test.name} ({assessment.test.id})"][1] += 1
        elif assessment.slot.slot == 'MID':
            tests_dict[assessment.test.education_center.name][f"{assessment.test.name} ({assessment.test.id})"][2] += 1
        else:
            tests_dict[assessment.test.education_center.name][f"{assessment.test.name} ({assessment.test.id})"][3] += 1 
        tests_dict[assessment.test.education_center.name]['count'] += 1
    
    return render(request, "vocational_guidance/dashboard.html", {
        'count_assessments': count_assessments,
        'count_unique_assessments': len(unique_assessments),
        'count_quota': int(count_assessments/2),
        'tests_dict': tests_dict,
        'slots_count': len(slots_count),
        "slots_count_link": len(slots_unique)
    })

@login_required(login_url='signin')
def profile(request, citizen_id):
    citizen = Citizen.objects.get(id=citizen_id)
    school = citizen.school
    bilet_distr = BiletDistribution.objects.filter(school=school)
    if len(bilet_distr) != 0:
        bilet_distr = bilet_distr[0]
        if bilet_distr.test_type == True:
            school_guid_type = "SPO"
        else:
            school_guid_type = "VO"
    else:
        school_guid_type = "VO"
    choosen_bundles = VocGuidTest.objects.filter(participants=citizen).values(
        "id", "name", "description", "img_link", "guid_type"
    )
    slots = TimeSlot.objects.all()
    slots_now = []
    for slot in slots:
        if date.today() < slot.date and slot.date < (date.today() + timedelta(days=7)):
            slots_now.append(slot)
    
    bundles = VocGuidTest.objects.filter(slots__in=slots_now, guid_type=school_guid_type).exclude(participants=citizen).values(
        "id", "name", "description", 
        "img_link", "guid_type", 
        "age_group", 'education_program_link', 
        'education_center__name', 'disability_types',
        "contact__full_name", "contact__email", "contact__phone"
    )

    group = VocGuidGroup.objects.filter(participants=citizen)
    choosen_type_presence = set()
    choosen_tests_dict = {}
    for guid_type in VocGuidTest.TYPE_CHOICES:
        choosen_tests_dict[guid_type[0]] = {}

    for bundle in choosen_bundles:
        if len(group) != 0:
            slot = TimeSlot.objects.filter(group__in=group)
            if len(slot) != 0:
                slot = slot[0]
            else:
                slot = None
        else:
            slot = None
        if slot == None:
            choosen_tests_dict[bundle["guid_type"]][bundle['id']] = {
                'id': bundle['id'],
                'name': bundle['name'],
                'description': bundle['description'],
                'slot': False
            }
        else:
            choosen_tests_dict[bundle["guid_type"]][bundle['id']] = {
                'id': bundle['id'],
                'name': bundle['name'],
                'description': bundle['description'],
                'slot': True,
                'date': slot.date,
                'time': slot.slot,
                'zoom_link': slot.zoom_link
            }
        choosen_type_presence.add(bundle["guid_type"])

    type_presence = set()
    tests_dict = {}
    for guid_type in VocGuidTest.TYPE_CHOICES:
        tests_dict[guid_type[0]] = {}
    if citizen.school_class.grade_number >= 10:
        age_group = '10-11'
    elif citizen.school_class.grade_number <= 7:
        age_group = '6-7'
    else:
        age_group = '8-9'

    citizen = Citizen.objects.get(email=request.user.email)
    for bundle in bundles:
        if bundle["age_group"] == "ALL" or bundle["age_group"] == age_group:
            if citizen.disability_type is None or citizen.disability_type.id == bundle['disability_types']:
                tests_dict[bundle["guid_type"]][bundle['id']] = {
                    'id': bundle['id'],
                    'name': bundle['name'],
                    'description': bundle['description'],
                    'img_link': bundle['img_link'],
                    'education_program_link': bundle['education_program_link'],
                    'education_center': bundle['education_center__name']
                }
                type_presence.add(bundle["guid_type"])
                

    message = ""
    if len(choosen_type_presence) == 0:
        message = "На данные момент Вы не записались не на одно из доступных профориентационных мероприятий. Вы можете просмотреть и выбрать интересующие вас их ниже."
    
    schools = School.objects.exclude(name=citizen.school.name)

    return render(request, "vocational_guidance/index.html",{
        'page_name': 'Личный кабинет',
        'user': citizen,
        'choosen_bundles': choosen_tests_dict,
        'choosen_type_presence': choosen_type_presence,
        'bundles': tests_dict,
        'slot': slot,
        'type_presence': type_presence,
        'message': message,
        "birthday": citizen.birthday.isoformat(),
        'schools': schools,
        "disability_types": DisabilityType.objects.all(),
        "choosed": len(choosen_bundles),
        "guid_type": school_guid_type
    })

@login_required(login_url='signin')
def school_dash(request, school_id):
    #Проверяем есть ли такая школа
    try:
        school = School.objects.get(id=school_id)
    except:
         return HttpResponseRedirect(reverse("index"))

    #Проверяем относиться ли пользователь к школе
    if request.user not in school.school_coordinators.all():
        return HttpResponseRedirect(reverse("index"))

    participants = Citizen.objects.filter(school=school)
    school_groups = VocGuidGroup.objects.filter(school=school)
    slots_enroll = TimeSlot.objects.filter(group__in=school_groups)
    groups_enroll = len(VocGuidGroup.objects.filter(school=school, slots__in=slots_enroll))
    tests = VocGuidTest.objects.all()
    quota = BiletDistribution.objects.filter(school=school).values("quota")[0]['quota']
    nonprofit_slots = TimeSlot.objects.filter(is_nonprofit=True)
    assessments = VocGuidAssessment.objects.filter(participant__in=participants)

    #Вычисляем лимит размера группы с учётом квоты школы
    school_limit = 0
    students_enroll = len(assessments.exclude(slot__in=nonprofit_slots))
    if quota != 0:
        limit = quota - students_enroll
        if limit >= 8:
            school_limit = 8
        else:
            school_limit = limit
    
    tests_dict = {}
    for test in tests:
        groups = school_groups.filter(bundle=test).annotate(participants_count=Count('participants'))
        if len(groups) != 0:
            tests_dict[f'{test.name} (Код пробы – {test.id})'] = {}
            #Добавляем контакты преподователя, если они существуют
            if TestContact.objects.filter(test=test).exists(): 
                tests_dict[f'{test.name} (Код пробы – {test.id})']['contact_name'] = test.contact.full_name
                tests_dict[f'{test.name} (Код пробы – {test.id})']['email'] = test.contact.email
                tests_dict[f'{test.name} (Код пробы – {test.id})']['phone'] = test.contact.phone
            else:
                tests_dict[f'{test.name} (Код пробы – {test.id})']['contact_name'] = None
            #Добавляем группы к тесту
            tests_dict[f'{test.name} (Код пробы – {test.id})']['groups'] = []
            for group in groups:
                group_list = []
                group_list.append(group)
                
                slots_list = []
                subscribe = None
                is_passed = False
                timeslot = TimeSlot.objects.filter(group=group, test=test)
                if len(timeslot) == 0:
                    #Определяем достаточно ли квоты для записи группы
                    if school_limit >= group.participants_count:
                        #Подбираем слоты по доступному количеству участников и датам
                        slot_date_limit = date.today() + timedelta(days=7)
                        slots = TimeSlot.objects.filter(
                            test=test,
                            date__gt = date.today(),
                            date__lte = slot_date_limit
                        )
                        for slot in slots:
                                slot_participants = VocGuidAssessment.objects.filter(slot=slot)
                                if len(slot_participants) + group.participants_count <= school_limit:
                                    slots_list.append([slot, len(slot_participants)])
                # Проверяем записанна ли группа на слот
                    group_list.append(None)
                else:
                    # Проверяем можно ли отказать от пробы (только если больше дня до начала)
                    group_list.append(timeslot[0])
                    if timeslot[0].date <= date.today():
                        subscribe = False
                        if timeslot[0].date < date.today():
                            is_passed = True
                    else:
                        subscribe = True
                        
                if len(slots_list) == 0:
                    slots_list =None
                group_list.append(slots_list)
                if subscribe is not None:
                    group_list.append(assessments.filter(slot=timeslot[0], participant__in=group.participants.all()))
                else:
                    group_list.append(group.participants.all())
                group_list.append(subscribe)
                group_list.append(is_passed)
                #Добавляем лист с данными по группе
                tests_dict[f'{test.name} (Код пробы – {test.id})']['groups'].append(group_list)
    
    #Статистика по зарегестрированным и записавщимся
    students = Citizen.objects.filter(school=school)
    school_stat = {}
    school_stat['students_count'] = len(students)
    school_stat['six_grade'] = len(students.filter(school_class__grade_number__in = [6,7]))
    school_stat['eight_grade'] = len(students.filter(school_class__grade_number__in = [8,9]))
    school_stat['ten_grade'] = len(students.filter(school_class__grade_number__in = [10,11]))

    school_stat['enroll_count'] = len(students.filter(voc_guid_tests__isnull=False))
    school_stat['six_grade_enroll']= len(students.filter(school_class__grade_number__in = [6,7], voc_guid_tests__isnull=False))
    school_stat['eight_grade_enroll'] = len(students.filter(school_class__grade_number__in = [8,9], voc_guid_tests__isnull=False))
    school_stat['ten_grade_enroll'] = len(students.filter(school_class__grade_number__in = [10,11], voc_guid_tests__isnull=False))

    school_stat['groups_count'] = len(school_groups)
    school_stat['groups_enroll'] = groups_enroll
    school_stat['school_limit'] = quota

    return render(request, 'vocational_guidance/school_dash.html', {
        'school': school,
        'school_stat': school_stat,
        'tests': tests_dict
    })

def tests_list(request):
    tests = {}
    for them in VocGuidTest.THEMES_CHOICES:
        tests[them[1]] = VocGuidTest.objects.filter(thematic_env=them[0])
    
    return render(request, 'vocational_guidance/tests_list.html', {
        "tests": tests
    })

def students_list(request, school_id):
    students = {}
    school = School.objects.filter(id=school_id)[0]
    classes = SchoolClass.objects.filter(school=school)
    for school_class in classes:
        class_students = Citizen.objects.filter(school_class=school_class)
        if len(class_students) > 0:
            students[school_class] = class_students
    
    return render(request, 'vocational_guidance/students_list.html', {
        "school": school,
        "classes": classes,
        "students": students,
        "disability_types": DisabilityType.objects.all(),
    })

def ed_center_dash(request, ed_center_id):
    pass

def region_dash(request):
    pass

@csrf_exempt
def choose_bundle(request):
    if request.method == "POST":
        bundle_id = request.POST["bundle_id"]
        citizen_id = request.POST["citizen"]
        test = VocGuidTest.objects.get(id=bundle_id)
        citizen = Citizen.objects.get(id=citizen_id)
        if citizen.school_class.grade_number >= 10:
            age_group = '10-11'
        elif citizen.school_class.grade_number <= 7:
            age_group = '6-7'
        else:
            age_group = '8-9'
        previous_bundles = VocGuidTest.objects.filter(
            participants=citizen_id,
            guid_type=test.guid_type
        )
        if len(previous_bundles) != 0:
            citizen.voc_guid_tests.remove(previous_bundles[0])
        test.participants.add(citizen_id)
        test.save()
        school = citizen.school
        
        groups = VocGuidGroup.objects.filter(bundle=test, school=school, age_group=age_group).annotate(participants_count=Count('participants'))
        
        if len(groups) == 0:
            create_group(citizen, test)
        else:
            add = False
            for group in groups:
                if group.participants_count < group.attendance_limit:
                    cheak_slot = TimeSlot.objects.filter(group=group)
                    if len(cheak_slot) == 0:
                        group.participants.add(citizen)
                        add = True
                        break
            if not add:
                create_group(citizen, test)
    return HttpResponseRedirect(reverse("index"))

def create_group(citizen, bundle):
    if citizen.school_class.grade_number >= 10:
        age_group = '10-11'
    elif citizen.school_class.grade_number <= 7:
        age_group = '6-7'
    else:
        age_group = '8-9'
    group = VocGuidGroup(
        school=citizen.school,
        bundle=bundle,
        attendance_limit=8,
        age_group=age_group 
    )
    group.save()
    group.participants.add(citizen)

@csrf_exempt
def reject_bundle(request):
    if request.method == "POST":
        bundle_id = request.POST["bundle_id"]
        citizen_id = request.POST["citizen"]
        bundle = VocGuidTest.objects.get(id=bundle_id)
        citizen = Citizen.objects.get(id=citizen_id)
        if citizen.school_class.grade_number >= 10:
            age_group = '10-11'
        elif citizen.school_class.grade_number <= 7:
            age_group = '6-7'
        else:
            age_group = '8-9'
        group = VocGuidGroup.objects.get(bundle=bundle,participants=citizen, age_group=age_group)

        citizen.voc_guid_tests.remove(bundle)
        citizen.voc_guid_groups.remove(group)
        citizen.save()
        if group.participants.count() == 0:
            group.delete()
    return HttpResponseRedirect(reverse("index"))

@csrf_exempt
def signin(request):
    message = None
    if request.user.is_authenticated:
        return HttpResponseRedirect(reverse("index"))
    elif request.method == "POST":
        email = request.POST["email"]
        password = request.POST["password"]
        citizen = Citizen.objects.filter(
            email=email,
        )
        user = authenticate(email=email, password=password)
        sch_group = Group.objects.filter(name='Школьник')
        if len(sch_group) != 0:
            sch_group = sch_group[0]
            if User.objects.filter(email=email, groups=sch_group):
                if user is not None:
                    if len(citizen) == 0:
                        message = "Ошибка авторизации, аккаунт не активирован. Обратитесь в поддержку – support@copp63.ru"
                    elif user is not None:
                        login(request, user)
                        return HttpResponseRedirect(reverse("index"))
                    else:
                        message = "Неверный логин и/или пароль."
    
    schools = School.objects.all()
    cities = set()
    for school in schools:
        cities.add(school.city) 
    return render(request, "vocational_guidance/login.html", {
        "message": message,
        "page_name": "Авторизация | Билет в будущее",
        'schools': schools,
        'cities': cities,
        'disability_types': DisabilityType.objects.all()
    }) 

def code_generator(size=6, chars=string.ascii_uppercase + string.digits):
    return ''.join(random.choice(chars) for _ in range(size))

@csrf_exempt
def password_recovery(request, step):
    if request.method == "POST":
        if step == 1:
            data = json.loads(request.body)
            email = data.get("email", "")
            user = User.objects.filter(email=email)
            if len(user) != 0:
                user = user[0]
                code = code_generator()
                user.code = code
                user.save()
                email = {
                    'subject': 'Востановление пароля copp63-coordinator.ru',
                    'html': f'<h1>Добрый день!</h1><p>Вы получили это письмо потому, что вы (либо кто-то, выдающий себя за вас) попросили выслать новый пароль к вашей учётной записи на сайте http://copp63-coordinator.ru/bilet/. <br> Если вы не просили выслать пароль, то не обращайте внимания на это письмо. <br> Код подтверждения для смены пароля: {code} <br> Это автоматическое письмо с на него не нужно отвечать.</p>',
                    'text': f'Добрый день!\n Вы получили это письмо потому, что вы (либо кто-то, выдающий себя за вас) попросили выслать новый пароль к вашей учётной записи на сайте http://copp63-coordinator.ru/bilet/. \n Если вы не просили выслать пароль, то не обращайте внимания на это письмо. \n Код подтверждения для смены пароля: {code} \n Это автоматическое письмо с на него не нужно отвечать.',
                    'from': {'name': 'ЦОПП СО', 'email': 'bvb@copp63.ru'},
                    'to': [
                        {'name': "f{user.first_name} {user.last_name}", 'email': email}
                    ],
                }
                SPApiProxy = mailing()
                SPApiProxy.smtp_send_mail(email)
                return JsonResponse({"message": "Email exict"}, status=201)
            return JsonResponse({"message": "Email not found"}, status=201)
        if step == 2:
            data = json.loads(request.body)
            email = data.get("email", "")
            code = data.get("code", "")
            user = User.objects.get(email=email)
            if user.code == code:
                return JsonResponse({"message": "Code matches"}, status=201)
            return JsonResponse({"message": "Code not matches"}, status=201)
        if step == 3:
            data = json.loads(request.body)
            email = data.get("email", "")
            password = data.get("password", "")
            confirmation = data.get("confirmation", "")
            if password == confirmation:
                user = User.objects.get(email=email)
                user.set_password(password)
                user.save()
                return JsonResponse({"message": "Password changed"}, status=201)
            return JsonResponse({"message": "Passwords mismatch"}, status=201)
    return HttpResponseRedirect(reverse("index"))

@login_required(login_url='signin')
@csrf_exempt
def change_password(request):
    if request.method == "POST":
        email = request.user.email
        current_password = request.POST["current_password"]
        password = request.POST["password"]
        confirmation = request.POST["confirmation"]
        user = authenticate(email=email, password=current_password)
        if user is not None:
            if password == confirmation:
                user.set_password(password)
                user.save()
                login(request, user)
    return HttpResponseRedirect(reverse("index"))

@csrf_exempt
def signup(request):
    if request.method == "POST":
        data = json.loads(request.body)
        email = data.get("email", "")
        password = data.get("password", "")
        confirmation = data.get("confirmation", "")

        phone = data.get("phone", "")
        first_name = data.get("first_name", "")
        last_name = data.get("last_name", "")
        middle_name = data.get("middle_name", "")
        birthday = data.get("birthday", "")
        disability_type = data.get("disability_type", "")

        school_id = data.get("school_id", "")
        grade_number = data.get("grade_number", "")
        grade_letter = data.get("grade_letter", "")
        school = School.objects.get(id=school_id)

        if disability_type != 'Выберите вид нарушения':
            disability_type = DisabilityType.objects.filter(id=disability_type)
            if len(disability_type) != 0:
                disability_type = disability_type[0]
        else:
            disability_type = None
        if password != confirmation:
            return JsonResponse({"message": "Password mismatch."}, status=201)
        try:
            user = User.objects.create_user(email, password)
            user.first_name = first_name
            user.last_name = last_name
            school_group = Group.objects.get(name='Школьник')
            user.groups.add(school_group)
            school_class = SchoolClass.objects.filter(
                school=school,
                grade_number=grade_number,
                grade_letter=grade_letter
            )
            if len(school_class) != 0:
                school_class = school_class[0]
            else:
                school_class = SchoolClass(
                    school=school,
                    grade_number=int(grade_number),
                    grade_letter=grade_letter.upper()
                )
                school_class.save()
            citizen = Citizen(
                first_name=first_name,
                last_name=last_name,
                middle_name=middle_name,
                birthday=birthday,
                email=email,
                social_status='SCHS',
                school=school,
                school_class=school_class,
                phone_number = phone,
                disability_type = disability_type
            )
            citizen.save()
            user.save()
        except IntegrityError:
            return JsonResponse({"message": "Email already taken."}, status=201)

        return JsonResponse({"message": "Account created successfully."}, status=201)
    return HttpResponseRedirect(reverse("index"))

def reg_choice(request):
    return HttpResponseRedirect(reverse("index"))

def reg_stage(request, choice, stage):
    return HttpResponseRedirect(reverse("index"))

@login_required(login_url='signin')
@csrf_exempt
def change_profile(request):
    if request.method == "POST":
        citizen_id = request.POST["id"]
        citizen = Citizen.objects.get(id=citizen_id)
        citizen.email = request.POST["email"]
        request.user.email = request.POST["email"]
        citizen.phone_number = request.POST["phone"]
        citizen.first_name = request.POST['name']
        request.user.first_name = request.POST["name"]
        citizen.last_name = request.POST['last_name']
        request.user.last_name = request.POST["last_name"]
        citizen.middle_name = request.POST['middle_name']
        citizen.birthday = request.POST['birthday']
        school_id = request.POST['school']
        school = School.objects.get(id=school_id)
        citizen.school = school
        grade_number = request.POST['school_class']
        grade_letter = request.POST['school_class_latter']
        try:
            disability_check = request.POST['disability-check']
        except:
            disability_check = False
        if disability_check != False:
            disability_type = request.POST['disability_type']
            citizen.disability_type = DisabilityType.objects.get(id=disability_type)
        else:
            citizen.disability_type = None
        
        school_class = SchoolClass.objects.filter(
                school=school,
                grade_number=grade_number,
                grade_letter=grade_letter
        )
        if len(school_class) != 0:
            school_class = school_class[0]
        else:
            school_class = SchoolClass(
                school=school,
                grade_number=int(grade_number),
                grade_letter=grade_letter.upper()
            )
            school_class.save()
        citizen.school_class = school_class
        request.user.save()
        citizen.save()
        return HttpResponseRedirect(reverse("index"))
    return HttpResponseRedirect(reverse("index"))

@login_required(login_url='bilet/login')
@csrf_exempt
def change_profile_teacher(request):
    if request.method == "POST":
        request.user.email = request.POST["email"]
        request.user.phone_number = request.POST["phone"]
        request.user.first_name = request.POST["name"]
        request.user.last_name = request.POST["last_name"]
        request.user.middle_name = request.POST['middle_name']
        request.user.save()
        school = School.objects.get(id=request.POST["school_id"])
        school.adress = request.POST['adress']
        school.save()
        return HttpResponseRedirect(reverse("index"))
    return HttpResponseRedirect(reverse("index"))

@login_required(login_url='signin')
@csrf_exempt
def change_profile_student(request):
    if request.method == "POST":
        citizen_id = request.POST["id"]
        citizen = Citizen.objects.get(id=citizen_id)
        email = request.POST["email"]
        citizen.email = request.POST["email"]
        user = User.objects.filter(email=request.POST["email_old"])
        if len(user) == 1:
            user = user[0]
            user.email = email
            user.first_name = request.POST["name"]
            user.last_name = request.POST["last_name"]
            user.save()
        citizen.phone_number = request.POST["phone"]
        citizen.first_name = request.POST['name']
        citizen.last_name = request.POST['last_name']
        citizen.middle_name = request.POST['middle_name']
        citizen.birthday = request.POST['birthday']
        
        grade_number = request.POST['school_class']
        grade_letter = request.POST['school_class_latter']
        
        disability_type = request.POST['disability_type']
        if disability_type != "Нет нарушений":
            citizen.disability_type = DisabilityType.objects.get(id=disability_type)
        else:
            citizen.disability_type = None
        school_id = request.POST['school']
        school = School.objects.get(id=school_id)
        school_class = SchoolClass.objects.filter(
                school=school,
                grade_number=grade_number,
                grade_letter=grade_letter
        )
        if len(school_class) != 0:
            school_class = school_class[0]
        else:
            school_class = SchoolClass(
                school=school,
                grade_number=int(grade_number),
                grade_letter=grade_letter.upper()
            )
            school_class.save()
        citizen.school_class = school_class
        citizen.save()
        return HttpResponseRedirect(reverse("students_list", args=(school.id,)))
    return HttpResponseRedirect(reverse("index"))

@login_required(login_url='signin')
def signout(request):
    if request.user.is_authenticated:
        logout(request)
    return HttpResponseRedirect(reverse("signin"))


@login_required
@csrf_exempt
def import_teachers(request):
    if request.method == "POST":
        form = ImportDataForm(request.POST, request.FILES)
        if form.is_valid():
            message = bvb_teachers(form)
        else:
            data = form.errors
        form = ImportDataForm()
        return render(request, "vocational_guidance/import_teachers.html",{
            'form': form,
            'message': message
        })
    else:
        form = ImportDataForm()
        return render(request, "vocational_guidance/import_teachers.html",{
            'form': form
        })

@login_required
@csrf_exempt
def import_slots(request):
    if request.method == "POST":
        form = ImportDataForm(request.POST, request.FILES)
        if form.is_valid():
            message = slots_import(form)
        else:
            data = form.errors
        form = ImportDataForm()
        return render(request, "vocational_guidance/import_slots.html",{
            'form': form,
            'message': message
        })
    else:
        form = ImportDataForm()
        return render(request, "vocational_guidance/import_slots.html",{
            'form': form
        })

@csrf_exempt
def import_bvb_matching(request):
    if request.method == "POST":
        form = ImportDataForm(request.POST, request.FILES)
        if form.is_valid():
            message = matching_bvb_students(form)
        else:
            data = form.errors
        form = ImportDataForm()
        return render(request, "vocational_guidance/bvb_matching.html",{
            'form': form,
            'message': message
        })
    else:
        form = ImportDataForm()
        return render(request, "vocational_guidance/bvb_matching.html",{
            'form': form
        })

@csrf_exempt
def choose_slot(request):
    if request.method == "POST":
        slot_id = request.POST["slot_id"]
        group_id = request.POST["group_id"]
        slot = TimeSlot.objects.get(id=slot_id)
        group = VocGuidGroup.objects.get(id=group_id)
        group.slots.clear()
        group.slots.add(slot)
        participants = Citizen.objects.filter(voc_guid_groups=group)
        for participant in participants:
            assessments = VocGuidAssessment.objects.filter(participant=participant, slot=slot)
            assessments.delete()
            assessment = VocGuidAssessment(
                participant=participant,
                test=slot.test,
                slot=slot
            )
            assessment.save()
    return HttpResponseRedirect(reverse("index"))

@csrf_exempt
def cancel_slot(request):
    if request.method == "POST":
        group_id = request.POST["group_id"]
        group = VocGuidGroup.objects.get(id=group_id)
        group.slots.clear()
        participants = Citizen.objects.filter(voc_guid_groups=group)
        for participant in participants:
            assessment = VocGuidAssessment.objects.filter(participant=participant)
            assessment.delete()
    return HttpResponseRedirect(reverse("index"))

def add_assessment_all(request):
    slots = TimeSlot.objects.all()
    for slot in slots:
        groups = VocGuidGroup.objects.filter(slots=slot)
        if len(groups) != 0:
            for group in groups:
                for participant in group.participants.all():
                    assessment = VocGuidAssessment.objects.filter(participant=participant, slot=slot)
                    if len(assessment) == 0:
                        assessment = VocGuidAssessment(
                            participant=participant,
                            test=slot.test,
                            slot=slot
                        )
                        assessment.save()
    return HttpResponseRedirect(reverse("index"))

def combine_groups(request):
    groups = VocGuidGroup.objects.filter(slots=None)
    for group in groups:
        duplicate_groups = VocGuidGroup.objects.filter(
            age_group=group.age_group, 
            bundle=group.bundle, 
            school=group.school
        )
        if len(duplicate_groups) != 0:
            participants = []
            for duplicate in duplicate_groups:
                for participant in duplicate.participants.all():
                    participants.append(participant)
                duplicate.delete()
            new_group = VocGuidGroup(
                age_group=group.age_group, 
                bundle=group.bundle, 
                school=group.school,
            )
            new_group.save()
            new_group.participants.set(participants)
            new_group.save()
            
    return HttpResponseRedirect(reverse("index"))

def regulate_groups(request):
    groups = VocGuidGroup.objects.filter(slots=None)
    for group in groups:
        group_count = len(group.participants.all())
        while group_count > 8:
            extra_participants_count = group_count - 8
            if extra_participants_count > 8:
                extra_participants_count = 8
            new_group = VocGuidGroup(
                age_group=group.age_group, 
                bundle=group.bundle, 
                school=group.school,
            )
            new_group.save()
            participant_count = 0
            for participant in group.participants.all():
                if participant_count < extra_participants_count:
                    participant.voc_guid_groups.clear()
                    participant.voc_guid_groups.add(new_group)
                    participant_count += 1
            group_count -= participant_count

    return HttpResponseRedirect(reverse("index"))

#Добавляем 0 квоту всем школам
def add_quotas_all(request):
    schools = School.objects.all()
    for school in schools:
        distribution = BiletDistribution.objects.filter(school=school)
        if len(distribution) == 0:
            distribution = BiletDistribution(
                school=school,
                test_type=False,
                quota=0
            )
        else:
            distribution = distribution[0]
            distribution.quota = 0
        distribution.save()
    return HttpResponseRedirect(reverse("index"))

def balance_quotas(request):
    if request.user.is_superuser:
        schools = School.objects.all()
        for school in schools:
            distribution = BiletDistribution.objects.filter(school=school)
            quota = BiletDistribution.objects.filter(school=school).values("quota")
            if len(quota) != 0:
                participants = Citizen.objects.filter(school=school)
                spent_quota = len(VocGuidAssessment.objects.filter(participant__in=participants, attendance=True))
                if quota[0]['quota'] < spent_quota:
                    distribution = distribution[0]
                    distribution.quota = spent_quota
                    distribution.save()
    return HttpResponseRedirect(reverse("quotas_dashboard"))

@csrf_exempt
def cancel_participant(request):
    if request.method == "POST":
        participant_id = request.POST["participant"]
        test = request.POST["test"]
        group = request.POST["group"]
        slots = request.POST["slots"]
        group = VocGuidGroup.objects.get(id=group)
        participant = Citizen.objects.get(id=participant_id)
        if participant.school_class.grade_number >= 10:
            age_group = '10-11'
        elif participant.school_class.grade_number <= 7:
            age_group = '6-7'
        else:
            age_group = '8-9'

        participant.voc_guid_groups.remove(group)
        if len(group.participants.all()) == 0:
            group.delete()

        school = participant.school
        groups = VocGuidGroup.objects.filter(bundle=test, school=school, age_group=age_group).annotate(participants_count=Count('participants'))
        
        test = VocGuidTest.objects.get(id=test)
        
        assessment = VocGuidAssessment.objects.filter(
            participant=participant, 
            test=test
        )
        assessment.delete()

        if len(groups) == 0:
            create_group(participant, test)
        else:
            for group in groups:
                if group.participants_count < group.attendance_limit:
                    cheak_slot = TimeSlot.objects.filter(group=group)
                    if len(cheak_slot) == 0:
                        group.participants.add(participant)
                        break
            else:
                create_group(participant, test)

            return HttpResponseRedirect(reverse("school_dash", args=(school.id,)))
    return HttpResponseRedirect(reverse("index"))

@login_required
@csrf_exempt
def import_external_slots(request):
    if EducationCenter.objects.filter(contact_person=request.user) != 0:
        education_center = EducationCenter.objects.filter(contact_person=request.user)[0]
        tests = VocGuidTest.objects.filter(education_center=education_center)
        schools = BiletDistribution.objects.all()
        if request.method == "POST":
            school_id = request.POST["school"]
            school = School.objects.get(id=school_id)
            test_id = request.POST["test"]
            test = VocGuidTest.objects.get(id=test_id)
            time = request.POST["time"]
            date = request.POST["date"]
            try:
                is_nonprofit = request.POST["is_nonprofit"]
                if is_nonprofit == "on":
                    is_nonprofit = True
            except:
                is_nonprofit = False
            report_link = request.POST["report_link"]
            slot = TimeSlot(
                test=test,
                date=date,
                slot=time,
                report_link=report_link,
                is_nonprofit=is_nonprofit
            )
            slot.save()
            group = VocGuidGroup(
                bundle=test,
                attendance_limit = 8,
                school=school,
                city=school.city
            )
            group.save()
            slot.group.add(group)
            slot.save()
            participants_count = 0
            for number in range(1,9):
                first_name = request.POST[f"name{number}"]
                middle_name = request.POST[f"middle_name{number}"]
                last_name = request.POST[f"last_name{number}"]
                grade_number = request.POST[f"school_class{number}"]
                grade_letter = request.POST[f"school_class_latter{number}"]
                if first_name != "" and last_name != "":
                    if grade_number != "" and grade_letter != "":
                        school_class = SchoolClass.objects.filter(
                            school=school,
                            grade_number=grade_number,
                            grade_letter=grade_letter.upper()
                        )
                        if len(school_class) != 0:
                            school_class = school_class[0]
                        else:
                            school_class = SchoolClass(
                                school=school,
                                grade_number=int(grade_number),
                                grade_letter=grade_letter.upper()
                            )
                            school_class.save()
                        citizen = Citizen(
                            first_name=first_name,
                            last_name=last_name,
                            middle_name=middle_name,
                            social_status='SCHS',
                            school=school,
                            school_class=school_class,
                        )
                        citizen.save()
                        group.participants.add(citizen)
                        group.save()
                        assessment = VocGuidAssessment(
                            participant=citizen,
                            test=test,
                            slot=slot,
                            attendance=True
                        )
                        assessment.save()
                        test.participants.add(citizen)
                        participants_count += 1

            return render(request, "vocational_guidance/import_external_slots.html", {
                "education_center": education_center,
                "tests": tests,
                "schools": schools,
                "test_limit": range(1,9),
                "participants_count": participants_count
            }) 
        else:
            return render(request, "vocational_guidance/import_external_slots.html", {
                "education_center": education_center,
                "tests": tests,
                "schools": schools,
                "test_limit": range(1,9)
            })


def mailing():
    REST_API_ID = 'e071900fe5ab9aa6dd4dec2f42160ead'
    REST_API_SECRET = '7e82daa1ccfd678487a894b3e3487967'
    TOKEN_STORAGE = 'memcached'
    MEMCACHED_HOST = '127.0.0.1:11211'
    SPApiProxy = PySendPulse(REST_API_ID, REST_API_SECRET, TOKEN_STORAGE, memcached_host=MEMCACHED_HOST)

    return SPApiProxy